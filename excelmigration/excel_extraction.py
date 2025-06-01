import json
import os
from pathlib import Path
from openpyxl import load_workbook

def normalize_url(url):
    """
    Normalize URLs by replacing relative paths with full network paths
    
    Args:
        url (str): The original URL
        
    Returns:
        str: The normalized URL
    """
    if not url:
        return url
    
    # Replace relative path with full network path
    if url.startswith("../.docs"):
        # Remove "../.docs" and replace with the full network path
        relative_part = url[8:]  # Remove "../.docs" (8 characters)
        # Convert forward slashes to backslashes for Windows network path consistency
        relative_part = relative_part.replace('/', '\\')
        normalized_url = f"file:///\\\\Dehesdna-a009a\\projekte\\k-z\\ofs\\Dokumentenservice\\TeileundStoffe{relative_part}"
        return normalized_url
    
    # If it's already a full file:// URL, return as is
    return url

def extract_excel_to_json_unified(excel_file_path, output_json_path=None, max_row=4608):
    """
    Extract data and hyperlinks from Excel file in a single pass
    
    Args:
        excel_file_path (str): Path to the Excel file
        output_json_path (str, optional): Path for output JSON file
        max_row (int): Maximum row to process (including header)
    
    Returns:
        dict: Extracted data as dictionary
    """
    try:
        print(f"Reading Excel file: {excel_file_path}")
        
        # Load workbook with openpyxl
        wb = load_workbook(excel_file_path, data_only=False)
        ws = wb.active
        
        # Get all column headers (row 1)
        headers = []
        for col in range(1, 28):  # Columns A to AA (1 to 27)
            cell = ws.cell(row=1, column=col)
            headers.append(cell.value if cell.value else f"Column_{chr(ord('A') + col - 1)}")
        
        print(f"Found {len(headers)} columns: {headers[:5]}{'...' if len(headers) > 5 else ''}")
        
        # Identify hyperlink columns (M to U = columns 13 to 21)
        hyperlink_col_indices = list(range(13, 22))  # M=13, N=14, ..., U=21
        hyperlink_col_names = [headers[i-1] for i in hyperlink_col_indices]
        print(f"Hyperlink columns: {hyperlink_col_names}")
        
        # Extract data row by row
        data_rows = []
        hyperlink_count = 0
        normalized_count = 0
        
        print(f"Processing rows 2 to {max_row}...")
        
        for row_num in range(2, max_row + 1):
            row_data = {}
            
            # Process each column
            for col_idx, header in enumerate(headers):
                col_num = col_idx + 1  # Convert to 1-based column number
                cell = ws.cell(row=row_num, column=col_num)
                
                # Check if this is a hyperlink column
                if col_num in hyperlink_col_indices:
                    if cell.hyperlink:
                        # Extract hyperlink information
                        target = cell.hyperlink.target
                        location = getattr(cell.hyperlink, 'location', None)
                        
                        # Handle different hyperlink formats
                        if target is None and location:
                            target = location
                        elif target and location:
                            target = f"{target}#{location}" if location else target
                        
                        # Normalize the URL
                        original_target = target
                        normalized_target = normalize_url(target)
                        
                        if original_target != normalized_target:
                            normalized_count += 1
                        
                        row_data[header] = {
                            'display_text': cell.value,
                            'url': normalized_target,
                            'original_url': original_target if original_target != normalized_target else None,
                            'location': location,
                            'tooltip': getattr(cell.hyperlink, 'tooltip', None)
                        }
                        hyperlink_count += 1
                        
                        # Debug: Show first few hyperlinks
                        if hyperlink_count <= 5:
                            print(f"  Row {row_num}, {header}: '{cell.value}' -> {normalized_target[:50]}...")
                        
                    elif cell.value:
                        # Check if the cell value looks like a file path
                        cell_value = str(cell.value).strip()
                        if (cell_value.startswith(('C:', 'D:', 'E:', '\\\\', './', '../')) or 
                            '\\' in cell_value or 
                            cell_value.endswith(('.pdf', '.doc', '.docx', '.xls', '.xlsx'))):
                            
                            # Normalize the inferred file path
                            normalized_path = normalize_url(cell_value)
                            
                            row_data[header] = {
                                'display_text': cell.value,
                                'url': normalized_path,
                                'original_url': cell_value if cell_value != normalized_path else None,
                                'location': None,
                                'tooltip': None,
                                'type': 'inferred_file_path'
                            }
                            
                            if cell_value != normalized_path:
                                normalized_count += 1
                        else:
                            # Regular text value in hyperlink column
                            row_data[header] = {
                                'display_text': cell.value,
                                'url': None,
                                'location': None,
                                'tooltip': None
                            }
                    else:
                        # Empty cell in hyperlink column
                        row_data[header] = None
                else:
                    # Regular data column
                    row_data[header] = cell.value
            
            data_rows.append(row_data)
            
            # Progress indicator
            if row_num % 500 == 0:
                print(f"  Processed {row_num - 1} rows...")
        
        print(f"Successfully processed {len(data_rows)} rows")
        print(f"Total hyperlinks found: {hyperlink_count}")
        print(f"URLs normalized: {normalized_count}")
        
        # Create the final data structure
        data_dict = {
            'metadata': {
                'total_rows': len(data_rows),
                'total_columns': len(headers),
                'columns': headers,
                'source_file': os.path.basename(excel_file_path),
                'hyperlinks_extracted': True,
                'hyperlink_columns': hyperlink_col_names,
                'url_normalization': {
                    'applied': True,
                    'rule': 'Replace "../.docs" with "file:///\\\\Dehesdna-a009a\\projekte\\k-z\\ofs\\Dokumentenservice\\TeileundStoffe"',
                    'normalized_count': normalized_count
                },
                'extraction_method': 'unified_openpyxl'
            },
            'data': data_rows
        }
        
        # Handle None values for JSON serialization
        def clean_none_values(obj):
            if isinstance(obj, dict):
                return {k: clean_none_values(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [clean_none_values(item) for item in obj]
            elif obj is None:
                return None
            else:
                return obj
        
        data_dict = clean_none_values(data_dict)
        
        # Determine output file path
        if output_json_path is None:
            excel_path = Path(excel_file_path)
            output_json_path = excel_path.parent / f"{excel_path.stem}.json"
        
        # Write to JSON file
        print(f"Writing JSON to: {output_json_path}")
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(data_dict, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"Successfully created JSON file with {len(data_rows)} rows")
        print(f"Output file: {output_json_path}")
        
        return data_dict
        
    except Exception as e:
        print(f"Error during extraction: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

def main():
    """Main function to run the unified extraction"""
    # Get the directory of this script
    script_dir = Path(__file__).parent
    excel_file = script_dir / "Verzeichnis.xlsx"
    
    if not excel_file.exists():
        print(f"Error: Excel file not found at {excel_file}")
        return
    
    try:
        # Extract data and hyperlinks in one pass
        data = extract_excel_to_json_unified(str(excel_file))
        
        # Print summary
        print("\n" + "="*60)
        print("UNIFIED EXTRACTION SUMMARY")
        print("="*60)
        print(f"Source file: {excel_file.name}")
        print(f"Total rows extracted: {data['metadata']['total_rows']}")
        print(f"Total columns: {data['metadata']['total_columns']}")
        print(f"Hyperlink columns: {', '.join(data['metadata']['hyperlink_columns'])}")
        print(f"URLs normalized: {data['metadata']['url_normalization']['normalized_count']}")
        print(f"Extraction method: {data['metadata']['extraction_method']}")
        
        # Show sample data
        if data['data']:
            print(f"\nFirst row sample:")
            first_row = data['data'][0]
            
            # Show regular columns
            regular_cols = ['Antrag-nummer', 'Teile-nummer', 'Freigabe']
            for col in regular_cols:
                if col in first_row:
                    print(f"  {col}: {first_row[col]}")
            
            # Show hyperlink columns
            print(f"\nHyperlink columns sample:")
            for col in data['metadata']['hyperlink_columns'][:3]:
                if col in first_row and first_row[col]:
                    if isinstance(first_row[col], dict):
                        print(f"  {col}: {first_row[col]['display_text']} -> {first_row[col]['url'][:50]}...")
                    else:
                        print(f"  {col}: {first_row[col]}")
                        
    except Exception as e:
        print(f"Failed to extract data: {str(e)}")

if __name__ == "__main__":
    main()
