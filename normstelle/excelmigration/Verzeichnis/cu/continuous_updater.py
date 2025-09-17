#!/usr/bin/env python3
"""
Continuous Excel Data Updater for Django App

Combines functionality from:
- btox.py: Fetches and converts Excel files
- excel_extraction_4.py: Extracts data with openpyxl
- url_cleanup.py: Cleans up URLs with replacement rules

This script continuously monitors and updates the Excel data for the Django app.
"""

from __future__ import annotations

import json
import os
import shutil
import sys
import time
import uuid
from pathlib import Path
from typing import Dict, List, Set, Tuple, Any, Optional
import traceback

# Import URL extraction functionality
try:
    from url_extract import extract_urls
except ImportError:
    # Fallback if import fails
    extract_urls = None


class ContinuousExcelUpdater:
    def __init__(self, django_base_path: str = None, test_mode: bool = False):
        """
        Initialize the continuous updater
        
        Args:
            django_base_path: Base path to the Django app (defaults to detected path)
            test_mode: If True, save all files to script directory instead of Django app
        """
        self.test_mode = test_mode
        self.script_dir = Path(__file__).resolve().parent
        
        if test_mode:
            # In test mode, use script directory for all operations
            self.temp_dir = self.script_dir / "temp"
            self.data_dir = self.script_dir
            self.django_base = self.script_dir
            self.log("TEST MODE: Files will be saved to script directory")
        else:
            # Auto-detect Django base path if not provided
            if django_base_path is None:
                # Assume script is in normstelle/excelmigration/Verzeichnis/cu/
                # and Django is at normie/normie/
                django_base_path = self.script_dir.parent.parent.parent.parent / "normie"
            
            self.django_base = Path(django_base_path)
            self.temp_dir = self.django_base / "normieapp" / "static" / "normieapp" / "temp"
            self.data_dir = self.django_base / "normieapp" / "static" / "normieapp" / "data"
        
        # Source paths for Excel files
        self.live_src = "\\\\deberdna-c010a\\GlobalDE\\DocumentManagement\\NormstelleShare\\TeileundStoffe\\Datei\\Verzeichnis.xlsb"
        self.test_src = "D:\\GlobalDE\\DocumentManagement\\NormstelleShare\\TeileundStoffe\\Datei\\Verzeichnis.xlsb"
        
        # Replacement rules for URL cleanup (loaded from replace file)
        self.replacement_rules = []
        self.target_replacement = ""
        self.ignore_patterns = set()
        self.dead_urls = set()
        
        # Load replacement rules from file
        self.load_replacement_rules()
        
        # Statistics
        self.stats = {
            'total_urls': 0,
            'fixed_urls': 0,
            'ignored_urls': 0,
            'unchanged_urls': 0,
            'error_urls': 0
        }
        
        # Ensure directories exist
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        self.data_dir.mkdir(parents=True, exist_ok=True)
        
        print(f"Initialized updater:")
        print(f"  Django base: {self.django_base}")
        print(f"  Temp dir: {self.temp_dir}")
        print(f"  Data dir: {self.data_dir}")

    def log(self, message: str):
        """Log message with timestamp"""
        timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
        print(f"[{timestamp}] {message}")

    def load_replacement_rules(self, replace_file: str = "replace"):
        """Load replacement rules from the replace file."""
        # Get the directory of this script for the replace file
        script_dir = Path(__file__).parent
        replace_path = script_dir / replace_file
        
        try:
            with open(replace_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Parse the replace file
            lines = content.strip().split('\n')
            current_section = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                    
                if line.lower() == 'replace:':
                    current_section = 'replace'
                elif line.lower() == 'with:':
                    current_section = 'with'
                elif line.lower() == 'ignore:':
                    current_section = 'ignore'
                elif current_section == 'replace':
                    # Add patterns to replace
                    self.replacement_rules.append(line)
                elif current_section == 'with':
                    # Set the target replacement
                    self.target_replacement = line
                elif current_section == 'ignore':
                    # Add patterns to ignore
                    self.ignore_patterns.add(line)
            
            self.log(f"Loaded {len(self.replacement_rules)} replacement rules from {replace_path}")
            self.log(f"Target replacement: {self.target_replacement}")
            self.log(f"Loaded {len(self.ignore_patterns)} ignore patterns")
            if self.dead_urls:
                self.log(f"Loaded {len(self.dead_urls)} dead URLs")
            
        except FileNotFoundError:
            self.log(f"Warning: Replace file '{replace_path}' not found, using defaults")
            # Fallback to basic rules if file not found
        except Exception as e:
            self.log(f"Error loading replacement rules: {e}")

    def pick_source_path(self) -> Path:
        """Return the first available source path (live first, then testing)."""
        live = Path(self.live_src)
        if live.exists():
            self.log(f"Using live source: {self.live_src}")
            return live
        
        test = Path(self.test_src)
        if test.exists():
            self.log(f"Using test source: {self.test_src}")
            return test
        
        raise FileNotFoundError(
            "Neither live nor testing source file was found.\n"
            f"Live: {self.live_src}\n"
            f"Test: {self.test_src}"
        )

    def copy_to_temp_dir(self, src_path: Path) -> Path:
        """Copy source file to temp directory with fixed name (reusable); return destination path."""
        # Use fixed filename - will be overwritten on each run to avoid accumulation
        dest_path = self.temp_dir / "Verzeichnis_current.xlsb"
        
        # Clean up existing file first
        if dest_path.exists():
            try:
                dest_path.unlink()
                self.log(f"Removed existing temp file: {dest_path}")
            except Exception as e:
                self.log(f"Warning: Could not remove existing temp file: {e}")
        
        self.log(f"Copying {src_path} to {dest_path}")
        shutil.copy2(src_path, dest_path)
        return dest_path

    def convert_xlsb_to_xlsx(self, xlsb_path: Path) -> Path:
        """Use Excel COM to convert .xlsb to .xlsx (overwrites destination)."""
        # Use fixed XLSX name (will be overwritten on each run)
        xlsx_path = self.temp_dir / "Verzeichnis_current.xlsx"
        
        # Clean up existing XLSX file first
        if xlsx_path.exists():
            try:
                xlsx_path.unlink()
                self.log(f"Removed existing temp XLSX: {xlsx_path}")
            except Exception as e:
                self.log(f"Warning: Could not remove existing temp XLSX: {e}")
        
        try:
            import win32com.client  # type: ignore
            import pythoncom
        except Exception as exc:
            raise RuntimeError(
                "pywin32 is required. Install with: pip install pywin32"
            ) from exc

        if xlsx_path.exists():
            xlsx_path.unlink()

        # Initialize COM for this thread
        pythoncom.CoInitialize()
        excel = None
        workbook = None
        try:
            self.log("Starting Excel COM for conversion...")
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            # UpdateLinks=0, ReadOnly=1 speeds up and prevents dialogs
            workbook = excel.Workbooks.Open(str(xlsb_path), UpdateLinks=0, ReadOnly=1)

            # 51 = xlOpenXMLWorkbook (.xlsx)
            workbook.SaveAs(str(xlsx_path), FileFormat=51)
            workbook.Close(SaveChanges=False)
            self.log(f"Converted to: {xlsx_path}")
            
        finally:
            try:
                if workbook is not None:
                    workbook.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            # Clean up COM
            pythoncom.CoUninitialize()
        
        return xlsx_path

    def normalize_url(self, url: str) -> str:
        """Pass-through for URLs during extraction - no normalization here to avoid conflicts with cleanup step"""
        # Don't normalize URLs during extraction - let the cleanup step handle all URL fixes
        # This prevents conflicts between extraction hardcoded rules and replace file rules
        return url if url else url

    def rgb_to_hex(self, rgb_value):
        """Convert RGB value to hex color code - same logic as working COM method"""
        if rgb_value is None:
            return None
        
        try:
            # Convert to integer if it's a float - exactly like the working version
            rgb_int = int(rgb_value)
            
            # Extract RGB components from the integer - same bit operations
            red = rgb_int & 255
            green = (rgb_int >> 8) & 255
            blue = (rgb_int >> 16) & 255
            
            return f"#{red:02X}{green:02X}{blue:02X}"
        
        except (ValueError, TypeError) as e:
            self.log(f"Warning: Could not convert RGB value {rgb_value} to hex: {e}")
            return None

    # Note: openpyxl_color_to_hex method removed - now using reliable Win32 COM for color extraction

    def map_color_to_status(self, color: str) -> str:
        """Map color codes to status descriptions"""
        color_mapping = {
            "#FFCC99": "not approved",
            "#CCFFCC": "approved", 
            "#CCFF99": "approved for first order",
            "#FFFFFF": "processing"
        }
        
        return color_mapping.get(color, "unknown")

    def extract_excel_to_json(self, excel_file_path: Path) -> Dict[str, Any]:
        """Extract data using optimized flow: COM for colors (reuse conversion session), openpyxl for data"""
        
        try:
            from openpyxl import load_workbook
            import win32com.client as win32
            import pythoncom
        except ImportError as e:
            raise ImportError(f"Required libraries missing: {e}. Install with: pip install openpyxl pywin32")
        
        start_time = time.time()
        max_row = 5000  # Limit for performance
        
        try:
            self.log("Starting optimized Excel extraction (COM for colors, openpyxl for data)")
            self.log(f"Reading Excel file: {excel_file_path}")
            
            # STEP 1: COM session for colors (reuse the conversion connection if this is XLSX already)
            color_start = time.time()
            colors = []
            color_count = 0
            
            self.log("Opening Excel with COM for color extraction...")
            
            # Initialize COM 
            pythoncom.CoInitialize()
            excel_app = None
            com_workbook = None
            
            try:
                excel_app = win32.Dispatch("Excel.Application")
                excel_app.Visible = False
                excel_app.DisplayAlerts = False
                
                # Open Excel file with COM
                com_workbook = excel_app.Workbooks.Open(str(excel_file_path), UpdateLinks=0, ReadOnly=1)
                com_worksheet = com_workbook.ActiveSheet
                
                # Get dimensions from COM
                actual_max_row = min(com_worksheet.UsedRange.Rows.Count, max_row)
                self.log(f"COM detected {actual_max_row} rows, extracting colors from column A...")
                
                # Extract colors while COM is open
                for row_num in range(2, actual_max_row + 1):
                    try:
                        cell_a = com_worksheet.Cells(row_num, 1)
                        rgb_value = cell_a.Interior.Color
                        hex_color = self.rgb_to_hex(rgb_value)
                        
                        if hex_color and hex_color != "#FFFFFF":
                            colors.append(hex_color)
                            color_count += 1
                            
                            # Debug: Show first few colors
                            if color_count <= 5:
                                self.log(f"  Row {row_num}: RGB={rgb_value} -> {hex_color}")
                        else:
                            colors.append("#FFFFFF")
                            
                    except Exception as e:
                        self.log(f"Warning: Could not extract color from row {row_num}: {e}")
                        colors.append("#FFFFFF")
                
            finally:
                # Clean up COM session
                if com_workbook:
                    com_workbook.Close(SaveChanges=False)
                if excel_app:
                    excel_app.Quit()
                pythoncom.CoUninitialize()
            
            color_time = time.time() - color_start
            self.log(f"Colors extracted with COM in {color_time:.3f}s ({color_count} colored cells found)")
            
            # STEP 2: Load workbook with openpyxl for fast data reading
            load_start = time.time()
            workbook = load_workbook(excel_file_path, data_only=False)
            worksheet = workbook.active
            
            load_time = time.time() - load_start
            self.log(f"Workbook loaded with openpyxl in {load_time:.3f}s")
            
            # Get worksheet dimensions
            actual_max_row = min(worksheet.max_row, max_row)
            max_col = min(worksheet.max_column, 27)  # Limit to AA (27 columns)
            
            self.log(f"Worksheet has {actual_max_row} rows and {max_col} columns")
            
            # Bulk data reading
            data_start = time.time()
            all_data = []
            for row in worksheet.iter_rows(min_row=1, max_row=actual_max_row, max_col=max_col, values_only=True):
                all_data.append(list(row))
            
            data_time = time.time() - data_start
            self.log(f"Bulk data read in {data_time:.3f}s")
            
            # Extract headers from first row
            headers = []
            if all_data:
                header_row = all_data[0]
                for col_idx in range(max_col):
                    if col_idx < len(header_row) and header_row[col_idx]:
                        headers.append(header_row[col_idx])
                    else:
                        headers.append(f"Column_{chr(ord('A') + col_idx)}")
            
            # Identify hyperlink columns (M to U = columns 13 to 21)
            hyperlink_col_indices = list(range(13, min(22, max_col + 1)))
            hyperlink_col_names = [headers[i-1] for i in hyperlink_col_indices if i <= len(headers)]
            self.log(f"Hyperlink columns: {hyperlink_col_names}")
            
            # Colors already extracted in STEP 1 (above) - no need to extract again
            
            # Extract hyperlinks
            hyperlink_start = time.time()
            hyperlink_map = {}
            hyperlink_count = 0
            
            for col_idx in hyperlink_col_indices:
                for row_num in range(2, actual_max_row + 1):
                    try:
                        cell = worksheet.cell(row=row_num, column=col_idx)
                        
                        if cell.hyperlink and cell.hyperlink.target:
                            target = cell.hyperlink.target
                            tooltip = getattr(cell.hyperlink, 'tooltip', None)
                            
                            # Apply URL normalization
                            normalized_target = self.normalize_url(target)
                            
                            hyperlink_map[(row_num, col_idx)] = {
                                'address': target,
                                'subaddress': None,
                                'full_target': normalized_target,
                                'screentip': tooltip
                            }
                            hyperlink_count += 1
                            
                    except Exception:
                        pass
            
            hyperlink_time = time.time() - hyperlink_start
            self.log(f"Hyperlinks extracted in {hyperlink_time:.3f}s ({hyperlink_count} found)")
            
            # Process data rows
            process_start = time.time()
            data_rows = []
            final_hyperlink_count = 0
            final_color_count = 0
            
            for row_idx in range(1, len(all_data)):
                row_num = row_idx + 1
                row_data_tuple = all_data[row_idx]
                row_data = {}
                
                # Extract color and status for this row
                color_idx = row_idx - 1
                if color_idx < len(colors):
                    cell_color = colors[color_idx]
                    if cell_color and cell_color != "#FFFFFF":
                        row_data['color'] = cell_color
                        row_data['status'] = self.map_color_to_status(cell_color)
                        final_color_count += 1
                    else:
                        row_data['color'] = "#FFFFFF"
                        row_data['status'] = "processing"
                else:
                    row_data['color'] = "#FFFFFF"
                    row_data['status'] = "processing"
                
                # Process each column for this row
                for col_idx, header in enumerate(headers):
                    col_num = col_idx + 1
                    
                    if col_num > max_col:
                        break
                    
                    # Get cell value from bulk data
                    cell_value = None
                    if row_data_tuple and col_idx < len(row_data_tuple):
                        cell_value = row_data_tuple[col_idx]
                        
                        # Handle datetime objects
                        if hasattr(cell_value, 'strftime'):
                            cell_value = cell_value.strftime('%Y-%m-%d %H:%M:%S+00:00')
                        
                        # Convert empty strings to None
                        elif isinstance(cell_value, str) and cell_value.strip() == '':
                            cell_value = None
                    
                    # Check if this is a hyperlink column
                    if col_num in hyperlink_col_indices:
                        hyperlink_data = hyperlink_map.get((row_num, col_num))
                        
                        if hyperlink_data:
                            target_url = hyperlink_data['full_target']
                            original_target = hyperlink_data['address']
                            
                            tooltip = hyperlink_data['screentip']
                            if tooltip is None:
                                tooltip = ''
                            
                            row_data[header] = {
                                'display_text': cell_value,
                                'url': target_url,
                                'original_url': None,  # No normalization during extraction
                                'tooltip': tooltip
                            }
                            final_hyperlink_count += 1
                            
                        elif cell_value:
                            # Check if the cell value looks like a file path
                            cell_value_str = str(cell_value).strip()
                            if (cell_value_str.startswith(('C:', 'D:', 'E:', '\\\\', './', '../')) or 
                                '\\' in cell_value_str or 
                                cell_value_str.endswith(('.pdf', '.doc', '.docx', '.xls', '.xlsx'))):
                                
                                # Store raw file path - no normalization during extraction
                                row_data[header] = {
                                    'display_text': cell_value,
                                    'url': cell_value_str,
                                    'original_url': None,
                                    'tooltip': None,
                                    'type': 'inferred_file_path'
                                }
                            else:
                                row_data[header] = {
                                    'display_text': cell_value,
                                    'url': None,
                                    'tooltip': None
                                }
                        else:
                            row_data[header] = None
                    else:
                        # Regular data column
                        if cell_value is None:
                            row_data[header] = None
                        elif isinstance(cell_value, str) and cell_value.strip() == '':
                            row_data[header] = None
                        else:
                            row_data[header] = cell_value
                
                data_rows.append(row_data)
                
                if row_idx % 1000 == 0:
                    self.log(f"  Processed {row_idx} rows...")
            
            process_time = time.time() - process_start
            total_processing_time = time.time() - start_time
            self.log(f"Data processing completed in {process_time:.3f}s")
            
            # Create the final data structure
            data_dict = {
                'metadata': {
                    'total_rows': len(data_rows),
                    'total_columns': len(headers),
                    'columns': headers,
                    'source_file': os.path.basename(excel_file_path),
                    'hyperlinks_extracted': True,
                    'hyperlink_columns': hyperlink_col_names,
                    'colors_extracted': True,
                    'color_mapping': {
                        '#FFCC99': 'not approved',
                        '#CCFFCC': 'approved',
                        '#CCFF99': 'approved for first order',
                        '#FFFFFF': 'processing'
                    },
                    'url_normalization': {
                        'applied': False,
                        'note': 'URL normalization disabled during extraction to prevent conflicts with cleanup step',
                        'normalized_count': 0
                    },
                    'extraction_method': 'optimized_hybrid_v2',
                    'color_extraction_method': 'win32_com_single_session',
                    'performance': {
                        'total_time': total_processing_time,
                        'load_time': load_time,
                        'data_read_time': data_time,
                        'color_extraction_time': color_time,
                        'hyperlink_extraction_time': hyperlink_time,
                        'data_processing_time': process_time,
                        'library_used': 'openpyxl',
                        'excel_app_needed': False
                    }
                },
                'data': data_rows
            }
            
            self.log(f"Successfully processed {len(data_rows)} rows")
            self.log(f"Total hyperlinks found: {final_hyperlink_count}")
            self.log(f"Colors extracted: {final_color_count} (single COM session)")
            self.log(f"Total time: {total_processing_time:.3f}s")
            self.log("Note: Optimized v2 - COM for colors (single session), openpyxl for data")
            self.log("Note: URL normalization disabled during extraction - will be handled in cleanup step")
            
            return data_dict
            
        except Exception as e:
            self.log(f"Error during Excel extraction: {str(e)}")
            traceback.print_exc()
            raise

    def should_ignore_url(self, url: str) -> bool:
        """Check if a URL should be ignored based on patterns and dead URL list."""
        if not url:
            return True
            
        # Check against ignore patterns
        for pattern in self.ignore_patterns:
            if pattern in url:
                return True
                
        # Check against dead URLs
        if url in self.dead_urls:
            return True
            
        # Check for HTTP/HTTPS URLs (should be ignored)
        if url.startswith(('http://', 'https://')):
            return True
            
        return False

    def fix_url(self, url: str) -> Tuple[str, bool]:
        """Fix a single URL based on replacement rules. Returns (fixed_url, was_changed)"""
        if not url or self.should_ignore_url(url):
            return url, False
            
        original_url = url
        fixed_url = url
        
        # Apply replacement rules
        for old_pattern in self.replacement_rules:
            if old_pattern in fixed_url:
                # Replace the old pattern with the new target
                before_fix = fixed_url
                fixed_url = fixed_url.replace(old_pattern, self.target_replacement)
                
                # Debug: Show rule application for first few fixes
                if self.stats.get('fixed_urls', 0) < 3:
                    self.log(f"    Applied rule '{old_pattern}' -> '{self.target_replacement}'")
                    self.log(f"    Before: {before_fix}")
                    self.log(f"    After:  {fixed_url}")
                
        # Normalize path separators (convert forward slashes to backslashes for Windows paths)
        if fixed_url.startswith('\\\\'):
            fixed_url = fixed_url.replace('/', '\\')
            
        return fixed_url, fixed_url != original_url

    def process_url_object(self, obj: Any) -> bool:
        """Process a URL object (dict with 'url' field) and fix its URL. Returns True if changed."""
        if not isinstance(obj, dict) or 'url' not in obj:
            return False
            
        original_url = obj['url']
        if not original_url:
            return False
            
        self.stats['total_urls'] += 1
        
        # Debug: Show first few URLs being processed
        if self.stats['total_urls'] <= 5:
            self.log(f"  Processing URL #{self.stats['total_urls']}: {original_url}")
        
        if self.should_ignore_url(original_url):
            self.stats['ignored_urls'] += 1
            if self.stats['ignored_urls'] <= 5:
                self.log(f"  -> Ignoring URL (matches ignore pattern): {original_url}")
            return False
            
        fixed_url, was_changed = self.fix_url(original_url)
        
        if was_changed:
            obj['url'] = fixed_url
            self.stats['fixed_urls'] += 1
            #self.log(f"  -> Fixed URL: {original_url} -> {fixed_url}")
            return True
        else:
            self.stats['unchanged_urls'] += 1
            if self.stats['unchanged_urls'] <= 5:
                self.log(f"  -> URL unchanged (no matching rules): {original_url}")
            return False

    def process_data_entry(self, entry: Dict, hyperlink_columns: List[str], debug_entry: bool = False) -> int:
        """Process a single data entry and fix all URLs within it. Returns number of changes."""
        changes = 0
        
        for column in hyperlink_columns:
            if column in entry and entry[column]:
                if debug_entry:
                    self.log(f"    Checking column '{column}': {type(entry[column])} = {str(entry[column])[:100]}...")
                
                if self.process_url_object(entry[column]):
                    changes += 1
                    if debug_entry:
                        self.log(f"    -> Fixed URL in column '{column}'")
            elif debug_entry:
                self.log(f"    Column '{column}': Not found or empty")
                    
        return changes

    def cleanup_urls_in_json(self, data: Dict[str, Any]) -> int:
        """Clean up URLs in the JSON data. Returns total number of changes."""
        self.log("Starting URL cleanup...")
        
        # Reset stats
        self.stats = {
            'total_urls': 0,
            'fixed_urls': 0,
            'ignored_urls': 0,
            'unchanged_urls': 0,
            'error_urls': 0
        }
        
        # Get hyperlink columns from metadata or use defaults
        hyperlink_columns = []
        if 'metadata' in data and 'hyperlink_columns' in data['metadata']:
            hyperlink_columns = data['metadata']['hyperlink_columns']
            self.log(f"Using hyperlink columns from metadata: {hyperlink_columns}")
        else:
            # Fallback to hardcoded list
            hyperlink_columns = [
                "Antrag", "Datenblatt", "Produkt-zulassung", "SDB MSDS",
                "Gefährdungsprüfungeurteilung", "Gefährdungsprüfung", 
                "Sonstiges", "Schriftverkehr", "Änd. Historie"
            ]
            self.log(f"Using fallback hyperlink columns: {hyperlink_columns}")
        
        # Debug: Show what we're looking for
        self.log(f"Replacement rules: {self.replacement_rules}")
        self.log(f"Target replacement: {self.target_replacement}")
        
        total_changes = 0
        
        # Process each data entry
        for i, entry in enumerate(data.get('data', [])):
            if i < 3:  # Debug first 3 entries
                self.log(f"DEBUG: Processing entry {i+1}")
                self.log(f"  Available columns: {list(entry.keys())}")
                
            changes = self.process_data_entry(entry, hyperlink_columns, debug_entry=i < 3)
            total_changes += changes
            
            if (i + 1) % 1000 == 0:
                self.log(f"  Processed {i + 1} entries...")
        
        # Update metadata
        if 'metadata' in data:
            data['metadata']['url_cleanup'] = {
                'applied': True,
                'total_changes': total_changes,
                'statistics': self.stats.copy(),
                'rules_applied': len(self.replacement_rules),
                'target_replacement': self.target_replacement,
                'hyperlink_columns_processed': hyperlink_columns
            }
        
        self.log(f"URL cleanup completed:")
        self.log(f"  Total URLs processed: {self.stats['total_urls']}")
        self.log(f"  URLs fixed: {self.stats['fixed_urls']}")
        self.log(f"  URLs ignored: {self.stats['ignored_urls']}")
        self.log(f"  URLs unchanged: {self.stats['unchanged_urls']}")
        self.log(f"  Total changes made: {total_changes}")
        
        return total_changes

    def save_json_data(self, data: Dict[str, Any], temp_path: Path, final_path: Path):
        """Save JSON data to temp location, then move to final location"""
        try:
            # Save to temp location first
            self.log(f"Saving JSON to temp location: {temp_path}")
            with open(temp_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            
            # Create backup of existing file if it exists
            if final_path.exists():
                backup_path = final_path.with_suffix('.json.backup')
                self.log(f"Creating backup: {backup_path}")
                shutil.copy2(final_path, backup_path)
            
            # Move temp file to final location
            self.log(f"Moving to final location: {final_path}")
            shutil.move(str(temp_path), str(final_path))
            
            
            self.log(f"Successfully saved JSON file: {final_path}")
            
        except Exception as e:
            self.log(f"Error saving JSON: {e}")
            raise

    def create_compressed_version(self, json_path: Path, compressed_path: Path):
        """Create compressed version of JSON for faster web loading"""
        try:
            # Import the optimizer class (inline to avoid import issues)
            class JSONOptimizer:
                def __init__(self):
                    self.base_url = "file:///\\\\deberdna-c010a\\GlobalDE\\DocumentManagement\\NormstelleShare\\TeileundStoffe\\\\"
                    self.column_map = {
                        "Antrag-nummer": "an", "Teile-nummer": "tn", "Freigabe": "fr",
                        "relevant für Luftfahrtteile": "rl", "Benennung": "bn",
                        "Produktname / Normkurzbezeichnung": "pn", "Produktzulassungs-spezifikation": "ps",
                        "Eingang": "ein", "Abschluss": "ab", "Abteilung": "abt", "Einsatzort": "eo",
                        "Antragsteller": "as", "Antrag": "ant", "Datenblatt": "db", "Produkt-zulassung": "pz",
                        "SDB MSDS": "sdb", "Gefährdungsprüfungeurteilung": "gpu", "Gefährdungsprüfung": "gp",
                        "Sonstiges": "so", "Schriftverkehr": "sv", "Änd. Historie": "ah", "Datum": "dt",
                        "Bearbeiter": "bb", "=CONCATENATE(\"Bemerkung \n(\",COUNTIF(C:C,\"TBD\"),\" offene Anträge)\")": "bem",
                        "GSK-Nr.": "gsk", "Lieferant": "lf", "Hersteller": "hs", "color": "c", "status": "s"
                    }
                    
                def compress_url(self, url: str) -> str:
                    if not url or not url.startswith(self.base_url):
                        return url
                    return url[len(self.base_url):]
                    
                def compress_document(self, doc):
                    if not doc or not isinstance(doc, dict):
                        return None
                    url = doc.get('url', '')
                    return self.compress_url(url) if url else None
                    
                def compress_data(self, data):
                    original_data = data.get('data', [])
                    compressed_data = []
                    for item in original_data:
                        compressed_item = {}
                        for orig_key, value in item.items():
                            short_key = self.column_map.get(orig_key, orig_key)
                            if value is None:
                                continue
                            if orig_key in ['Antrag', 'Datenblatt', 'Produkt-zulassung', 'SDB MSDS',
                                          'Gefährdungsprüfungeurteilung', 'Gefährdungsprüfung', 
                                          'Sonstiges', 'Schriftverkehr', 'Änd. Historie']:
                                compressed_doc = self.compress_document(value)
                                if compressed_doc:
                                    compressed_item[short_key] = compressed_doc
                            else:
                                compressed_item[short_key] = value
                        compressed_data.append(compressed_item)
                    
                    return {
                        "metadata": {
                            "total_rows": len(compressed_data),
                            "base_url": self.base_url,
                            "column_map": {v: k for k, v in self.column_map.items()},
                            "compressed": True,
                            "version": "1.0"
                        },
                        "data": compressed_data
                    }
            
            # Create compressed version
            self.log(f"Creating compressed version: {compressed_path}")
            
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            optimizer = JSONOptimizer()
            compressed_data = optimizer.compress_data(data)
            
            with open(compressed_path, 'w', encoding='utf-8') as f:
                json.dump(compressed_data, f, separators=(',', ':'))
            
            # Get file sizes for logging
            original_size = json_path.stat().st_size / (1024 * 1024)
            compressed_size = compressed_path.stat().st_size / (1024 * 1024)
            savings = (1 - compressed_size / original_size) * 100
            
            self.log(f"Compression complete: {original_size:.1f}MB → {compressed_size:.1f}MB ({savings:.1f}% smaller)")
            
        except Exception as e:
            self.log(f"Warning: Failed to create compressed version: {e}")


    def extract_urls_from_json(self, json_file_path: Path, output_file_name: str = None) -> bool:
        """Extract URLs from a JSON file and save to text file
        
        Args:
            json_file_path: Path to the JSON file to extract URLs from
            output_file_name: Name of the output file (defaults to urls_[json_filename].txt)
            
        Returns:
            bool: True if extraction was successful, False otherwise
        """
        if extract_urls is None:
            self.log("Warning: URL extraction not available (url_extract module not found)")
            return False
            
        if not json_file_path.exists():
            self.log(f"Warning: JSON file not found for URL extraction: {json_file_path}")
            return False
            
        try:
            # Determine output file name
            if output_file_name is None:
                json_stem = json_file_path.stem
                output_file_name = f"urls_{json_stem}.txt"
            
            # Create full output path in same directory as JSON file
            output_file_path = json_file_path.parent / output_file_name
            
            self.log(f"Extracting URLs from: {json_file_path}")
            self.log(f"Output file: {output_file_path}")
            
            # Save current working directory
            original_cwd = os.getcwd()
            
            try:
                # Use the existing temp directory instead of creating new ones
                temp_extraction_dir = self.temp_dir
                
                # Change to the temp directory
                os.chdir(temp_extraction_dir)
                
                # Copy the JSON file to the expected name in temp directory
                temp_json_path = temp_extraction_dir / "Verzeichnis.json"
                if temp_json_path.exists():
                    temp_json_path.unlink()  # Remove existing copy
                shutil.copy2(json_file_path, temp_json_path)
                self.log(f"Created temporary copy in temp dir: {temp_json_path}")
                
                # Extract URLs (unique URLs by default) - output will be in temp directory
                temp_output_file = str(output_file_name)
                urls = extract_urls(extract_all=False, output_file=temp_output_file)
                
                # Move the output file to the final location
                temp_output_path = temp_extraction_dir / temp_output_file
                if temp_output_path.exists():
                    shutil.move(str(temp_output_path), str(output_file_path))
                    self.log(f"Successfully extracted {len(urls) if urls else 0} unique URLs to {output_file_path}")
                else:
                    self.log(f"Warning: Output file not created: {temp_output_path}")
                    return False
                
                # Clean up the temporary JSON copy
                try:
                    if temp_json_path.exists():
                        temp_json_path.unlink()
                        self.log(f"Cleaned up temporary JSON copy: {temp_json_path}")
                except Exception as cleanup_error:
                    self.log(f"Warning: Could not clean up temp JSON: {cleanup_error}")
                
                return True
                
            finally:
                # Restore original working directory
                os.chdir(original_cwd)
                
        except Exception as e:
            self.log(f"Error during URL extraction from {json_file_path}: {e}")
            traceback.print_exc()
            return False

    def cleanup_temp_files(self):
        """Clean up temporary files (optimized for single-file approach)"""
        try:
            # Clean up current temp files and any leftover timestamped files
            patterns = [
                "Verzeichnis_current.xlsb",    # Current XLSB file
                "Verzeichnis_current.xlsx",    # Current XLSX file
                "Verzeichnis_*.xlsb",          # Legacy timestamped XLSB files
                "Verzeichnis_*.xlsx",          # Legacy timestamped XLSX files  
                "Verzeichnis.xlsb",            # Legacy fixed name
                "Verzeichnis.xlsx",            # Legacy fixed name
                "Verzeichnis_original_temp.json",
                "Verzeichnis_temp.json"
            ]
            
            files_deleted = 0
            for pattern in patterns:
                for temp_file in self.temp_dir.glob(pattern):
                    try:
                        temp_file.unlink()
                        files_deleted += 1
                        self.log(f"Deleted temp file: {temp_file}")
                    except Exception as file_error:
                        self.log(f"Warning: Could not delete {temp_file}: {file_error}")
            
            if files_deleted == 0:
                self.log("No temp files to clean up")
            else:
                self.log(f"Cleaned up {files_deleted} temp files")
                    
        except Exception as e:
            self.log(f"Warning: Could not delete temp files: {e}")
    
    def cleanup_single_file(self, file_path: Path):
        """Clean up a single temp file immediately after use"""
        try:
            if file_path.exists():
                file_path.unlink()
                self.log(f"Immediately cleaned up: {file_path}")
        except Exception as e:
            self.log(f"Warning: Could not clean up {file_path}: {e}")
            
    def cleanup_legacy_testing_files(self):
        """Clean up legacy testing files that are no longer needed"""
        try:
            # Files that are no longer created (legacy testing files)
            legacy_patterns = [
                "Verzeichnis_original.json",      # Original extracted data
                "Verzeichnis_original.json.backup", # Backup of original
                "urls_original.txt",              # URL list from original
                "urls_cleaned.txt"                # URL list from cleaned
            ]
            
            files_deleted = 0
            for pattern in legacy_patterns:
                legacy_file = self.data_dir / pattern
                if legacy_file.exists():
                    try:
                        legacy_file.unlink()
                        files_deleted += 1
                        self.log(f"Removed legacy testing file: {legacy_file}")
                    except Exception as file_error:
                        self.log(f"Warning: Could not remove legacy file {legacy_file}: {file_error}")
            
            if files_deleted > 0:
                self.log(f"Cleaned up {files_deleted} legacy testing files")
                    
        except Exception as e:
            self.log(f"Warning: Could not clean up legacy files: {e}")

    def run_single_update(self) -> bool:
        """Run a single update cycle. Returns True if successful."""
        try:
            self.log("="*60)
            self.log("STARTING UPDATE CYCLE")
            self.log("="*60)
            
            # Step 1: Fetch source Excel file
            src_path = self.pick_source_path()
            
            # Step 2: Copy to temp directory (with fixed filename)
            temp_xlsb = self.copy_to_temp_dir(src_path)
            self.log(f"Created temp file: {temp_xlsb.name}")
            
            try:
                # Step 3: Convert XLSB to XLSX (with fixed filename)
                temp_xlsx = self.convert_xlsb_to_xlsx(temp_xlsb)
                self.log(f"Converting to XLSX: {temp_xlsx.name}")
                
                # Clean up XLSB immediately after conversion
                self.cleanup_single_file(temp_xlsb)
                
                try:
                    # Step 4: Extract data to JSON
                    json_data = self.extract_excel_to_json(temp_xlsx)
                    
                    # Clean up XLSX immediately after extraction
                    self.cleanup_single_file(temp_xlsx)
                    
                except Exception as extract_error:
                    # Clean up XLSX even if extraction fails
                    self.cleanup_single_file(temp_xlsx)
                    raise extract_error
                    
            except Exception as conversion_error:
                # Clean up XLSB even if conversion fails
                self.cleanup_single_file(temp_xlsb)
                raise conversion_error
            
            # Step 5: Cleanup URLs in the main data (skip saving original)
            changes_made = self.cleanup_urls_in_json(json_data)
            self.log(f"URL cleanup completed with {changes_made} changes")
            
            # Step 6: Save cleaned JSON to final location (only essential file)
            temp_json = self.temp_dir / "Verzeichnis_temp.json"
            final_json = self.data_dir / "Verzeichnis.json"
            self.save_json_data(json_data, temp_json, final_json)
            
            # Clean up temp JSON file immediately after save
            self.cleanup_single_file(temp_json)
            
            # Step 7.5: Create compressed version for web interface
            try:
                compressed_json = self.data_dir / "Verzeichnis_compressed.json"
                self.create_compressed_version(final_json, compressed_json)
            except Exception as e:
                self.log(f"Warning: Failed to create compressed version: {e}")
            
            # Step 7: Skip URL extraction (not needed for production)
            
            # Step 8: Clean up legacy testing files (no longer needed)
            self.cleanup_legacy_testing_files()
            
            # Step 9: Final cleanup (any remaining temp files)
            self.cleanup_temp_files()
            
            self.log("="*60)
            self.log("UPDATE CYCLE COMPLETED SUCCESSFULLY")
            self.log(f"Main file: {final_json}")
            self.log(f"Compressed file: {compressed_json}")
            self.log(f"Backup file: {final_json.with_suffix('.json.backup')}")
            self.log(f"Total rows: {json_data['metadata']['total_rows']}")
            self.log(f"URL changes: {changes_made}")
            self.log("="*60)
            
            return True
            
        except Exception as e:
            self.log(f"ERROR in update cycle: {e}")
            traceback.print_exc()
            
            # Try to cleanup temp files even on error
            try:
                self.cleanup_temp_files()
            except:
                pass
                
            return False

    def run_continuous(self, interval_minutes: int = 30):
        """Run continuous updates with specified interval"""
        self.log(f"Starting continuous updates (every {interval_minutes} minutes)")
        self.log("Press Ctrl+C to stop")
        
        try:
            while True:
                success = self.run_single_update()
                
                if success:
                    self.log(f"Waiting {interval_minutes} minutes until next update...")
                else:
                    self.log(f"Update failed, waiting {interval_minutes} minutes before retry...")
                
                # Wait for the specified interval
                time.sleep(interval_minutes * 60)
                
        except KeyboardInterrupt:
            self.log("Continuous updates stopped by user")
        except Exception as e:
            self.log(f"Continuous updates stopped due to error: {e}")
            traceback.print_exc()


def main():
    """Main function"""
    print("Excel Continuous Updater for Django App")
    print("="*50)
    
    # Parse command line arguments
    test_mode = False
    run_mode = "once"  # Default mode
    interval = 30  # Default interval for continuous mode
    
    args = sys.argv[1:]
    i = 0
    while i < len(args):
        arg = args[i]
        
        if arg == "--test":
            test_mode = True
            print("TEST MODE: Files will be saved to script directory")
        elif arg == "--once":
            run_mode = "once"
        elif arg == "--continuous":
            run_mode = "continuous"
            # Check if next argument is an interval
            if i + 1 < len(args) and args[i + 1].isdigit():
                i += 1
                try:
                    interval = int(args[i])
                except ValueError:
                    print("Invalid interval, using default 30 minutes")
        elif arg.startswith("--"):
            print(f"Unknown option: {arg}")
            print("Usage:")
            print("  python continuous_updater.py [--test] --once")
            print("  python continuous_updater.py [--test] --continuous [interval_minutes]")
            print("  python continuous_updater.py [--test]  (defaults to --once)")
            print("")
            print("Options:")
            print("  --test                    Save all files to script directory instead of Django app")
            print("  --once                    Run single update cycle")
            print("  --continuous [interval]   Run continuous updates (default: 30 minutes)")
            return 1
        else:
            # Check if it's a number for continuous mode interval
            if run_mode == "continuous" and arg.isdigit():
                try:
                    interval = int(arg)
                except ValueError:
                    print(f"Invalid argument: {arg}")
                    return 1
            else:
                print(f"Invalid argument: {arg}")
                print("Usage:")
                print("  python continuous_updater.py [--test] --once")
                print("  python continuous_updater.py [--test] --continuous [interval_minutes]")
                print("  python continuous_updater.py [--test]  (defaults to --once)")
                return 1
        
        i += 1
    
    # Initialize updater with test mode
    try:
        updater = ContinuousExcelUpdater(test_mode=test_mode)
    except Exception as e:
        print(f"Failed to initialize updater: {e}")
        return 1
    
    # Run based on mode
    if run_mode == "once":
        # Run single update
        success = updater.run_single_update()
        return 0 if success else 1
    elif run_mode == "continuous":
        # Run continuous updates
        updater.run_continuous(interval)
        return 0
    else:
        # Should not reach here
        print("Internal error: Invalid run mode")
        return 1


if __name__ == "__main__":
    sys.exit(main())
