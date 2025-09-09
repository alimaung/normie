#!/usr/bin/env python3
"""
Color Comparison Test: Win32 COM vs openpyxl
Compare color extraction methods on specific rows where we know there are discrepancies
"""

import os
import sys
import time
from pathlib import Path
import traceback

# Add the script directory to the path for imports
script_dir = Path(__file__).parent
sys.path.insert(0, str(script_dir))

def old_com_rgb_to_hex(rgb_value):
    """Original working function from excel_extraction.py"""
    if rgb_value is None:
        return None
    
    try:
        # Convert to integer if it's a float
        rgb_int = int(rgb_value)
        
        # Extract RGB components from the integer
        red = rgb_int & 255
        green = (rgb_int >> 8) & 255
        blue = (rgb_int >> 16) & 255
        
        return f"#{red:02X}{green:02X}{blue:02X}"
    
    except (ValueError, TypeError) as e:
        print(f"Warning: Could not convert RGB value {rgb_value} to hex: {e}")
        return None

def new_openpyxl_color_to_hex(color_obj):
    """New function from continuous_updater.py with current mappings"""
    if not color_obj:
        return "#FFFFFF"
    
    try:
        # Handle openpyxl indexed colors
        if hasattr(color_obj, 'indexed') and color_obj.indexed is not None:
            # Current mappings from continuous_updater.py
            indexed_to_rgb_map = {
                42: 13434828,  # Light green -> should be #CCFFCC (approved)
                43: 10079487,  # Should be #FFCC99 based on Win32 COM  
                47: 10079164,  # Should be different from 43
            }
            
            indexed_val = color_obj.indexed
            if indexed_val in indexed_to_rgb_map:
                # Use the old working RGB-to-hex conversion logic
                rgb_int = indexed_to_rgb_map[indexed_val]
                red = rgb_int & 255
                green = (rgb_int >> 8) & 255
                blue = (rgb_int >> 16) & 255
                hex_color = f"#{red:02X}{green:02X}{blue:02X}"
                return hex_color
            
            # Unknown indexed color
            return "#FFFFFF"
        
        # Handle RGB colors
        elif hasattr(color_obj, 'rgb') and color_obj.rgb:
            rgb_value = color_obj.rgb
            if isinstance(rgb_value, str):
                if len(rgb_value) == 8:
                    return f"#{rgb_value[2:].upper()}"
                elif len(rgb_value) == 6:
                    return f"#{rgb_value.upper()}"
            return "#FFFFFF"
        
        else:
            return "#FFFFFF"
            
    except Exception as e:
        print(f"Warning: Could not convert color object {type(color_obj)} to hex: {e}")
        return "#FFFFFF"

def test_row_colors():
    """Test color extraction on specific rows with both methods"""
    
    # Find the Excel file
    script_dir = Path(__file__).parent
    xlsx_file = script_dir / "Verzeichnis.xlsx"
    
    if not xlsx_file.exists():
        print(f"Error: Excel file not found at {xlsx_file}")
        print("Please ensure Verzeichnis.xlsx exists in the script directory")
        return
    
    print("="*70)
    print("COLOR EXTRACTION COMPARISON TEST")
    print("="*70)
    print(f"Excel file: {xlsx_file}")
    print(f"Target column: A (column 1)")
    print()
    
    # Test rows - you mentioned we should use the same cells we targeted before
    # Based on our previous analysis, let's test a range around the problematic row
    test_rows = []
    
    # Find the row with Antrag-nummer "141/2025" - let's test around that area
    # For now, let's test a few specific rows that are likely to have colors
    for row_num in range(4615, 4619):  # Your suggested range
        test_rows.append(row_num)
    
    # Also test some earlier rows that might have colors
    for row_num in range(2, 6):  # First few data rows
        test_rows.append(row_num)
    
    print("Testing rows:", test_rows)
    print()
    
    # Test with Win32 COM method
    print("1. WIN32 COM METHOD (OLD - WORKING)")
    print("-" * 40)
    
    try:
        import win32com.client as win32
        import pythoncom
        
        pythoncom.CoInitialize()
        excel_app = win32.Dispatch("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False
        
        workbook = excel_app.Workbooks.Open(os.path.abspath(xlsx_file))
        worksheet = workbook.ActiveSheet
        
        com_results = {}
        
        for row_num in test_rows:
            try:
                cell_a = worksheet.Cells(row_num, 1)
                rgb_value = cell_a.Interior.Color
                hex_color = old_com_rgb_to_hex(rgb_value)
                
                # Get cell value for context
                cell_value = cell_a.Value
                
                com_results[row_num] = {
                    'rgb_int': rgb_value,
                    'hex_color': hex_color,
                    'cell_value': cell_value
                }
                
                print(f"Row {row_num:4d}: RGB={rgb_value:>10} -> {hex_color} | Cell: {str(cell_value)[:30]}")
                
            except Exception as e:
                print(f"Row {row_num:4d}: ERROR - {e}")
                com_results[row_num] = {'error': str(e)}
        
        workbook.Close(SaveChanges=False)
        excel_app.Quit()
        pythoncom.CoUninitialize()
        
    except Exception as e:
        print(f"Error with COM method: {e}")
        com_results = {}
    
    print()
    
    # Test with openpyxl method
    print("2. OPENPYXL METHOD (NEW - PROBLEMATIC)")
    print("-" * 40)
    
    try:
        from openpyxl import load_workbook
        
        workbook = load_workbook(xlsx_file, data_only=False)
        worksheet = workbook.active
        
        openpyxl_results = {}
        
        for row_num in test_rows:
            try:
                cell = worksheet.cell(row=row_num, column=1)
                
                # Get color object details
                color_obj = cell.fill.start_color if cell.fill else None
                hex_color = new_openpyxl_color_to_hex(color_obj)
                
                # Get cell value for context
                cell_value = cell.value
                
                # Get indexed color info if available
                indexed_info = "None"
                rgb_info = "None"
                if color_obj:
                    try:
                        if hasattr(color_obj, 'indexed') and color_obj.indexed is not None:
                            indexed_info = f"indexed={color_obj.indexed}"
                    except Exception as e:
                        indexed_info = f"indexed=ERROR: {e}"
                    
                    try:
                        if hasattr(color_obj, 'rgb') and color_obj.rgb:
                            rgb_info = f"rgb={color_obj.rgb}"
                    except Exception as e:
                        rgb_info = f"rgb=ERROR: {e}"
                    
                    # Also check the raw color object
                    if indexed_info == "None" and rgb_info == "None":
                        indexed_info = f"type={type(color_obj).__name__}"
                
                openpyxl_results[row_num] = {
                    'hex_color': hex_color,
                    'cell_value': cell_value,
                    'indexed_info': indexed_info,
                    'rgb_info': rgb_info
                }
                
                print(f"Row {row_num:4d}: {indexed_info:>15} -> {hex_color} | Cell: {str(cell_value)[:30]}")
                
            except Exception as e:
                print(f"Row {row_num:4d}: ERROR - {e}")
                openpyxl_results[row_num] = {'error': str(e)}
        
    except Exception as e:
        print(f"Error with openpyxl method: {e}")
        openpyxl_results = {}
    
    print()
    
    # Compare results
    print("3. COMPARISON")
    print("-" * 40)
    
    print(f"{'Row':>4} | {'COM RGB':>10} | {'COM Hex':>8} | {'openpyxl':>15} | {'openpyxl Hex':>8} | {'Match?':>6}")
    print("-" * 70)
    
    for row_num in test_rows:
        com_result = com_results.get(row_num, {})
        openpyxl_result = openpyxl_results.get(row_num, {})
        
        com_rgb = com_result.get('rgb_int', 'ERROR')
        com_hex = com_result.get('hex_color', 'ERROR')
        openpyxl_indexed = openpyxl_result.get('indexed_info', 'ERROR')
        openpyxl_hex = openpyxl_result.get('hex_color', 'ERROR')
        
        match = "✓" if com_hex == openpyxl_hex else "✗"
        
        print(f"{row_num:4d} | {str(com_rgb):>10} | {str(com_hex):>8} | {str(openpyxl_indexed):>15} | {str(openpyxl_hex):>8} | {match:>6}")
    
    print()
    print("="*70)
    print("Analysis:")
    print("- Look for rows where COM and openpyxl return different hex colors")
    print("- Check the indexed color values that openpyxl finds")
    print("- Update the indexed_to_rgb_map in continuous_updater.py accordingly")
    print("="*70)

if __name__ == "__main__":
    test_row_colors()
