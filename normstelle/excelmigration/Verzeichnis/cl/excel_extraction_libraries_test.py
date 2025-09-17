import json
import os
import time
from pathlib import Path
from typing import Dict, List, Any, Optional

def normalize_url(url):
    """
    Normalize URLs by replacing relative paths with full network paths
    """
    if not url:
        return url
    
    if url.startswith("../.docs"):
        relative_part = url[8:]
        relative_part = relative_part.replace('/', '\\')
        normalized_url = f"file:///\\\\Dehesdna-a009a\\projekte\\k-z\\ofs\\Dokumentenservice\\TeileundStoffe{relative_part}"
        return normalized_url
    
    return url

def rgb_to_hex(rgb_value):
    """Convert RGB value to hex color code"""
    if rgb_value is None:
        return None
    
    try:
        rgb_int = int(rgb_value)
        red = rgb_int & 255
        green = (rgb_int >> 8) & 255
        blue = (rgb_int >> 16) & 255
        return f"#{red:02X}{green:02X}{blue:02X}"
    except (ValueError, TypeError) as e:
        print(f"Warning: Could not convert RGB value {rgb_value} to hex: {e}")
        return None

def map_color_to_status(color):
    """Map color codes to status descriptions"""
    color_mapping = {
        "#FFCC99": "not approved",
        "#CCFFCC": "approved", 
        "#CCFF99": "approved for first order",
        "#FFFFFF": "processing"
    }
    return color_mapping.get(color, "unknown")

# =============================================================================
# 1. XLWINGS APPROACH
# =============================================================================

def extract_with_xlwings(excel_file_path: str) -> Optional[Dict]:
    """
    Extract using xlwings - modern Excel API
    Often 2-5x faster than pywin32
    """
    try:
        import xlwings as xw
    except ImportError:
        print("xlwings not installed. Install with: pip install xlwings")
        return None
    
    start_time = time.time()
    print(f"\n[{time.strftime('%H:%M:%S')}] ===== XLWINGS EXTRACTION =====")
    
    app = None
    wb = None
    
    try:
        # Create Excel app
        app = xw.App(visible=False, add_book=False)
        wb = app.books.open(excel_file_path)
        ws = wb.sheets[0]
        
        # Get worksheet dimensions
        used_range = ws.used_range
        max_row = used_range.last_cell.row
        max_col = min(used_range.last_cell.column, 27)
        
        print(f"[{time.strftime('%H:%M:%S')}] Worksheet: {max_row} rows, {max_col} columns")
        
        # BULK READ ALL DATA
        data_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Bulk reading all data...")
        
        # Read all data in one shot
        all_data = ws.range(f'A1:{chr(64+max_col)}{max_row}').value
        
        print(f"[{time.strftime('%H:%M:%S')}] Data read in {time.time() - data_start:.3f}s")
        
        # BULK READ COLORS
        color_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Reading colors...")
        
        # xlwings can read colors in bulk
        color_range = ws.range(f'A2:A{max_row}')
        colors = []
        
        # Try bulk color reading
        try:
            color_values = color_range.color
            if isinstance(color_values, list):
                colors = [rgb_to_hex(color) if color else "#FFFFFF" for color in color_values]
            else:
                # Single color or different format
                colors = [rgb_to_hex(color_values) if color_values else "#FFFFFF"] * (max_row - 1)
        except:
            # Fallback to individual color reading
            for row in range(2, max_row + 1):
                try:
                    color = ws.range(f'A{row}').color
                    colors.append(rgb_to_hex(color) if color else "#FFFFFF")
                except:
                    colors.append("#FFFFFF")
        
        print(f"[{time.strftime('%H:%M:%S')}] Colors read in {time.time() - color_start:.3f}s")
        
        # HYPERLINKS (xlwings may have better support)
        hyperlink_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Reading hyperlinks...")
        
        hyperlink_map = {}
        hyperlink_col_indices = list(range(13, min(22, max_col + 1)))
        
        for col_idx in hyperlink_col_indices:
            col_letter = chr(64 + col_idx)
            for row in range(2, max_row + 1):
                try:
                    cell = ws.range(f'{col_letter}{row}')
                    if hasattr(cell, 'hyperlink') and cell.hyperlink:
                        hyperlink_map[(row, col_idx)] = {
                            'address': cell.hyperlink,
                            'full_target': cell.hyperlink
                        }
                except:
                    pass
        
        print(f"[{time.strftime('%H:%M:%S')}] Hyperlinks read in {time.time() - hyperlink_start:.3f}s")
        print(f"[{time.strftime('%H:%M:%S')}] Found {len(hyperlink_map)} hyperlinks")
        
        total_time = time.time() - start_time
        print(f"[{time.strftime('%H:%M:%S')}] XLWINGS TOTAL TIME: {total_time:.3f}s")
        
        return {
            'method': 'xlwings',
            'total_time': total_time,
            'data': all_data,
            'colors': colors,
            'hyperlinks': hyperlink_map,
            'rows': max_row,
            'cols': max_col
        }
        
    except Exception as e:
        print(f"xlwings extraction failed: {e}")
        return None
        
    finally:
        if wb:
            wb.close()
        if app:
            app.quit()

# =============================================================================
# 2. OPENPYXL APPROACH
# =============================================================================

def extract_with_openpyxl(excel_file_path: str) -> Optional[Dict]:
    """
    Extract using openpyxl - pure Python, no Excel needed
    Often 10-50x faster for reading
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        print("openpyxl not installed. Install with: pip install openpyxl")
        return None
    
    start_time = time.time()
    print(f"\n[{time.strftime('%H:%M:%S')}] ===== OPENPYXL EXTRACTION =====")
    
    try:
        # Load workbook
        load_start = time.time()
        wb = load_workbook(excel_file_path, data_only=False)  # data_only=False to get formulas/hyperlinks
        ws = wb.active
        
        print(f"[{time.strftime('%H:%M:%S')}] Workbook loaded in {time.time() - load_start:.3f}s")
        
        # Get dimensions
        max_row = min(ws.max_row, 5000)
        max_col = min(ws.max_column, 27)
        
        print(f"[{time.strftime('%H:%M:%S')}] Worksheet: {max_row} rows, {max_col} columns")
        
        # BULK READ ALL DATA
        data_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Bulk reading all data...")
        
        # openpyxl can read ranges very efficiently
        all_data = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            all_data.append(list(row))
        
        print(f"[{time.strftime('%H:%M:%S')}] Data read in {time.time() - data_start:.3f}s")
        
        # BULK READ COLORS
        color_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Reading colors...")
        
        colors = []
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=1)
            
            # Get fill color
            try:
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    # openpyxl uses ARGB, we need RGB
                    argb = cell.fill.start_color.rgb
                    if len(argb) == 8:  # ARGB format
                        rgb_hex = f"#{argb[2:]}"  # Remove alpha channel
                    else:
                        rgb_hex = f"#{argb}"
                    colors.append(rgb_hex)
                else:
                    colors.append("#FFFFFF")
            except:
                colors.append("#FFFFFF")
        
        print(f"[{time.strftime('%H:%M:%S')}] Colors read in {time.time() - color_start:.3f}s")
        
        # HYPERLINKS (openpyxl has excellent hyperlink support)
        hyperlink_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Reading hyperlinks...")
        
        hyperlink_map = {}
        hyperlink_col_indices = list(range(13, min(22, max_col + 1)))
        
        for col_idx in hyperlink_col_indices:
            for row in range(2, max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                
                if cell.hyperlink:
                    hyperlink_map[(row, col_idx)] = {
                        'address': cell.hyperlink.target,
                        'full_target': cell.hyperlink.target,
                        'tooltip': getattr(cell.hyperlink, 'tooltip', None)
                    }
        
        print(f"[{time.strftime('%H:%M:%S')}] Hyperlinks read in {time.time() - hyperlink_start:.3f}s")
        print(f"[{time.strftime('%H:%M:%S')}] Found {len(hyperlink_map)} hyperlinks")
        
        total_time = time.time() - start_time
        print(f"[{time.strftime('%H:%M:%S')}] OPENPYXL TOTAL TIME: {total_time:.3f}s")
        
        return {
            'method': 'openpyxl',
            'total_time': total_time,
            'data': all_data,
            'colors': colors,
            'hyperlinks': hyperlink_map,
            'rows': max_row,
            'cols': max_col
        }
        
    except Exception as e:
        print(f"openpyxl extraction failed: {e}")
        import traceback
        traceback.print_exc()
        return None

# =============================================================================
# 3. PANDAS APPROACH
# =============================================================================

def extract_with_pandas(excel_file_path: str) -> Optional[Dict]:
    """
    Extract using pandas - excellent for bulk data operations
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
    except ImportError:
        print("pandas not installed. Install with: pip install pandas openpyxl")
        return None
    
    start_time = time.time()
    print(f"\n[{time.strftime('%H:%M:%S')}] ===== PANDAS EXTRACTION =====")
    
    try:
        # BULK READ DATA with pandas (super fast)
        data_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Bulk reading data with pandas...")
        
        # Read Excel with pandas - this is extremely fast
        df = pd.read_excel(excel_file_path, engine='openpyxl', header=0, nrows=4999)
        
        print(f"[{time.strftime('%H:%M:%S')}] Data read in {time.time() - data_start:.3f}s")
        print(f"[{time.strftime('%H:%M:%S')}] DataFrame: {df.shape[0]} rows, {df.shape[1]} columns")
        
        # Convert to our format
        all_data = [df.columns.tolist()]  # Headers
        all_data.extend(df.values.tolist())  # Data rows
        
        # For hyperlinks and colors, we still need openpyxl
        wb = load_workbook(excel_file_path, data_only=False)
        ws = wb.active
        
        max_row = min(ws.max_row, 5000)
        max_col = min(ws.max_column, 27)
        
        # Colors (fast with openpyxl)
        color_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Reading colors...")
        
        colors = []
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=1)
            try:
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    argb = cell.fill.start_color.rgb
                    rgb_hex = f"#{argb[2:]}" if len(argb) == 8 else f"#{argb}"
                    colors.append(rgb_hex)
                else:
                    colors.append("#FFFFFF")
            except:
                colors.append("#FFFFFF")
        
        print(f"[{time.strftime('%H:%M:%S')}] Colors read in {time.time() - color_start:.3f}s")
        
        # Hyperlinks
        hyperlink_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Reading hyperlinks...")
        
        hyperlink_map = {}
        hyperlink_col_indices = list(range(13, min(22, max_col + 1)))
        
        for col_idx in hyperlink_col_indices:
            for row in range(2, max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.hyperlink:
                    hyperlink_map[(row, col_idx)] = {
                        'address': cell.hyperlink.target,
                        'full_target': cell.hyperlink.target
                    }
        
        print(f"[{time.strftime('%H:%M:%S')}] Hyperlinks read in {time.time() - hyperlink_start:.3f}s")
        print(f"[{time.strftime('%H:%M:%S')}] Found {len(hyperlink_map)} hyperlinks")
        
        total_time = time.time() - start_time
        print(f"[{time.strftime('%H:%M:%S')}] PANDAS TOTAL TIME: {total_time:.3f}s")
        
        return {
            'method': 'pandas',
            'total_time': total_time,
            'data': all_data,
            'colors': colors,
            'hyperlinks': hyperlink_map,
            'rows': max_row,
            'cols': max_col
        }
        
    except Exception as e:
        print(f"pandas extraction failed: {e}")
        import traceback
        traceback.print_exc()
        return None

# =============================================================================
# 4. POLARS APPROACH (Ultra-fast)
# =============================================================================

def extract_with_polars(excel_file_path: str) -> Optional[Dict]:
    """
    Extract using polars - ultra-fast dataframe library
    """
    try:
        import polars as pl
        from openpyxl import load_workbook
    except ImportError:
        print("polars not installed. Install with: pip install polars openpyxl")
        return None
    
    start_time = time.time()
    print(f"\n[{time.strftime('%H:%M:%S')}] ===== POLARS EXTRACTION =====")
    
    try:
        # BULK READ DATA with polars (blazing fast)
        data_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Bulk reading data with polars...")
        
        # Read Excel with polars - this should be very fast
        df = pl.read_excel(excel_file_path, sheet_name=0)
        
        print(f"[{time.strftime('%H:%M:%S')}] Data read in {time.time() - data_start:.3f}s")
        print(f"[{time.strftime('%H:%M:%S')}] DataFrame: {df.shape[0]} rows, {df.shape[1]} columns")
        
        # Convert to our format
        all_data = [df.columns]  # Headers
        all_data.extend(df.to_numpy().tolist())  # Data rows
        
        # For hyperlinks and colors, use openpyxl (as polars doesn't support these)
        wb = load_workbook(excel_file_path, data_only=False)
        ws = wb.active
        
        max_row = min(ws.max_row, 5000)
        max_col = min(ws.max_column, 27)
        
        # Colors and hyperlinks (same as pandas approach)
        color_start = time.time()
        colors = []
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=1)
            try:
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    argb = cell.fill.start_color.rgb
                    rgb_hex = f"#{argb[2:]}" if len(argb) == 8 else f"#{argb}"
                    colors.append(rgb_hex)
                else:
                    colors.append("#FFFFFF")
            except:
                colors.append("#FFFFFF")
        
        print(f"[{time.strftime('%H:%M:%S')}] Colors read in {time.time() - color_start:.3f}s")
        
        hyperlink_start = time.time()
        hyperlink_map = {}
        hyperlink_col_indices = list(range(13, min(22, max_col + 1)))
        
        for col_idx in hyperlink_col_indices:
            for row in range(2, max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.hyperlink:
                    hyperlink_map[(row, col_idx)] = {
                        'address': cell.hyperlink.target,
                        'full_target': cell.hyperlink.target
                    }
        
        print(f"[{time.strftime('%H:%M:%S')}] Hyperlinks read in {time.time() - hyperlink_start:.3f}s")
        print(f"[{time.strftime('%H:%M:%S')}] Found {len(hyperlink_map)} hyperlinks")
        
        total_time = time.time() - start_time
        print(f"[{time.strftime('%H:%M:%S')}] POLARS TOTAL TIME: {total_time:.3f}s")
        
        return {
            'method': 'polars',
            'total_time': total_time,
            'data': all_data,
            'colors': colors,
            'hyperlinks': hyperlink_map,
            'rows': max_row,
            'cols': max_col
        }
        
    except Exception as e:
        print(f"polars extraction failed: {e}")
        return None

# =============================================================================
# MAIN COMPARISON
# =============================================================================

def main():
    """Run comprehensive library comparison"""
    script_dir = Path(__file__).parent
    excel_file = script_dir / "Verzeichnis.xlsx"
    
    if not excel_file.exists():
        print(f"Error: Excel file not found at {excel_file}")
        return
    
    print("="*80)
    print("EXCEL LIBRARY PERFORMANCE COMPARISON")
    print("="*80)
    print(f"Testing file: {excel_file}")
    
    results = []
    
    # Test all libraries
    libraries = [
        ("xlwings", extract_with_xlwings),
        ("openpyxl", extract_with_openpyxl),
        ("pandas", extract_with_pandas),
        ("polars", extract_with_polars)
    ]
    
    for name, func in libraries:
        print(f"\n{'='*20} TESTING {name.upper()} {'='*20}")
        result = func(str(excel_file))
        if result:
            results.append(result)
    
    # Compare results
    print("\n" + "="*80)
    print("PERFORMANCE COMPARISON SUMMARY")
    print("="*80)
    
    if results:
        # Sort by performance
        results.sort(key=lambda x: x['total_time'])
        
        print(f"{'Library':<15} {'Time (s)':<10} {'Speedup':<10} {'Hyperlinks':<12} {'Status'}")
        print("-" * 70)
        
        baseline_time = 154.7  # v2 baseline
        
        for result in results:
            speedup = baseline_time / result['total_time']
            print(f"{result['method']:<15} {result['total_time']:<10.3f} {speedup:<10.2f}x {len(result['hyperlinks']):<12} ✅")
        
        # Show baseline for comparison
        print(f"{'pywin32 (v2)':<15} {baseline_time:<10.1f} {'1.00':<10}x {'17182':<12} (baseline)")
        
        fastest = results[0]
        print(f"\n🏆 WINNER: {fastest['method'].upper()} - {fastest['total_time']:.3f}s")
        print(f"🚀 SPEEDUP vs pywin32: {baseline_time / fastest['total_time']:.2f}x faster")
        
    else:
        print("❌ No libraries succeeded. Install required packages:")
        print("  pip install xlwings openpyxl pandas polars")

if __name__ == "__main__":
    main()
