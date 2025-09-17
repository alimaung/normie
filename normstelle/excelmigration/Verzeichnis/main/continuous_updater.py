#!/usr/bin/env python3
"""
Continuous Excel Data Updater for Django App

Combines functionality from:
- btox.py: Fetches and converts Excel files
- excel_extraction_4.py: Extracts data with openpyxl
- url_cleanup.py: Cleans up URLs with replacement rules

This script continuously monitors and updates the Excel data for the Django app.
"""

from __future__ import annotations

import json
import os
import shutil
import sys
import time
from pathlib import Path
from typing import Dict, List, Set, Tuple, Any, Optional
import traceback


class ContinuousExcelUpdater:
    def __init__(self, django_base_path: str = None):
        """
        Initialize the continuous updater
        
        Args:
            django_base_path: Base path to the Django app (defaults to detected path)
        """
        # Auto-detect Django base path if not provided
        if django_base_path is None:
            # Assume script is in normstelle/excelmigration/Verzeichnis/main/
            # and Django is at normie/normie/
            script_dir = Path(__file__).resolve().parent
            django_base_path = script_dir.parent.parent.parent.parent / "normie"
        
        self.django_base = Path(django_base_path)
        self.temp_dir = self.django_base / "normieapp" / "static" / "normieapp" / "temp"
        self.data_dir = self.django_base / "normieapp" / "static" / "normieapp" / "data"
        
        # Source paths for Excel files
        self.live_src = "\\\\deberdna-c010a\\GlobalDE\\DocumentManagement\\Ofs\\obl\\Dokumentenservice\\TeileundStoffe\\Datei\\Verzeichnis.xlsb"
        self.test_src = "D:\\GlobalDE\\DocumentManagement\\Ofs\\obl\\Dokumentenservice\\TeileundStoffe\\Datei\\Verzeichnis.xlsb"
        
        # Replacement rules for URL cleanup (loaded from replace file)
        self.replacement_rules = []
        self.target_replacement = ""
        self.ignore_patterns = set()
        self.dead_urls = set()
        
        # Load replacement rules from file
        self.load_replacement_rules()
        
        # Statistics
        self.stats = {
            'total_urls': 0,
            'fixed_urls': 0,
            'ignored_urls': 0,
            'unchanged_urls': 0,
            'error_urls': 0
        }
        
        # Ensure directories exist
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        self.data_dir.mkdir(parents=True, exist_ok=True)
        
        print(f"Initialized updater:")
        print(f"  Django base: {self.django_base}")
        print(f"  Temp dir: {self.temp_dir}")
        print(f"  Data dir: {self.data_dir}")

    def log(self, message: str):
        """Log message with timestamp"""
        timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
        print(f"[{timestamp}] {message}")

    def load_replacement_rules(self, replace_file: str = "replace"):
        """Load replacement rules from the replace file."""
        # Get the directory of this script for the replace file
        script_dir = Path(__file__).parent
        replace_path = script_dir / replace_file
        
        try:
            with open(replace_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Parse the replace file
            lines = content.strip().split('\n')
            current_section = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                    
                if line.lower() == 'replace:':
                    current_section = 'replace'
                elif line.lower() == 'with:':
                    current_section = 'with'
                elif line.lower() == 'ignore:':
                    current_section = 'ignore'
                elif current_section == 'replace':
                    # Add patterns to replace
                    self.replacement_rules.append(line)
                elif current_section == 'with':
                    # Set the target replacement
                    self.target_replacement = line
                elif current_section == 'ignore':
                    # Add patterns to ignore
                    self.ignore_patterns.add(line)
            
            self.log(f"Loaded {len(self.replacement_rules)} replacement rules from {replace_path}")
            self.log(f"Target replacement: {self.target_replacement}")
            self.log(f"Loaded {len(self.ignore_patterns)} ignore patterns")
            if self.dead_urls:
                self.log(f"Loaded {len(self.dead_urls)} dead URLs")
            
        except FileNotFoundError:
            self.log(f"Warning: Replace file '{replace_path}' not found, using defaults")
            # Fallback to basic rules if file not found
        except Exception as e:
            self.log(f"Error loading replacement rules: {e}")

    def pick_source_path(self) -> Path:
        """Return the first available source path (live first, then testing)."""
        live = Path(self.live_src)
        if live.exists():
            self.log(f"Using live source: {self.live_src}")
            return live
        
        test = Path(self.test_src)
        if test.exists():
            self.log(f"Using test source: {self.test_src}")
            return test
        
        raise FileNotFoundError(
            "Neither live nor testing source file was found.\n"
            f"Live: {self.live_src}\n"
            f"Test: {self.test_src}"
        )

    def copy_to_temp_dir(self, src_path: Path) -> Path:
        """Copy source file to temp directory as Verzeichnis.xlsb; return destination path."""
        dest_path = self.temp_dir / "Verzeichnis.xlsb"
        self.log(f"Copying {src_path} to {dest_path}")
        shutil.copy2(src_path, dest_path)
        return dest_path

    def convert_xlsb_to_xlsx(self, xlsb_path: Path) -> Path:
        """Use Excel COM to convert .xlsb to .xlsx (overwrites destination)."""
        xlsx_path = self.temp_dir / "Verzeichnis.xlsx"
        
        try:
            import win32com.client  # type: ignore
        except Exception as exc:
            raise RuntimeError(
                "pywin32 is required. Install with: pip install pywin32"
            ) from exc

        if xlsx_path.exists():
            xlsx_path.unlink()

        excel = None
        workbook = None
        try:
            self.log("Starting Excel COM for conversion...")
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            # UpdateLinks=0, ReadOnly=1 speeds up and prevents dialogs
            workbook = excel.Workbooks.Open(str(xlsb_path), UpdateLinks=0, ReadOnly=1)

            # 51 = xlOpenXMLWorkbook (.xlsx)
            workbook.SaveAs(str(xlsx_path), FileFormat=51)
            workbook.Close(SaveChanges=False)
            self.log(f"Converted to: {xlsx_path}")
            
        finally:
            try:
                if workbook is not None:
                    workbook.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
        
        return xlsx_path

    def normalize_url(self, url: str) -> str:
        """Pass-through for URLs during extraction - no normalization here to avoid conflicts with cleanup step"""
        # Don't normalize URLs during extraction - let the cleanup step handle all URL fixes
        # This prevents conflicts between extraction hardcoded rules and replace file rules
        return url if url else url

    def rgb_to_hex(self, rgb_value):
        """Convert RGB value to hex color code"""
        if rgb_value is None:
            return None
        
        try:
            # Convert to integer if it's a float
            rgb_int = int(rgb_value)
            
            # Extract RGB components from the integer
            red = rgb_int & 255
            green = (rgb_int >> 8) & 255
            blue = (rgb_int >> 16) & 255
            
            return f"#{red:02X}{green:02X}{blue:02X}"
        
        except (ValueError, TypeError) as e:
            self.log(f"Warning: Could not convert RGB value {rgb_value} to hex: {e}")
            return None

    def openpyxl_color_to_hex(self, color_obj):
        """Convert openpyxl color object to hex color code"""
        if not color_obj:
            return "#FFFFFF"
        
        try:
            # Handle openpyxl indexed colors
            if hasattr(color_obj, 'indexed') and color_obj.indexed is not None:
                indexed_color_map = {
                    42: "#CCFFCC",  # Light green (approved)
                    43: "#CCFF99",  # Light green-yellow (approved for first order)  
                    47: "#FFCC99",  # Light orange (not approved)
                }
                
                indexed_val = color_obj.indexed
                if indexed_val in indexed_color_map:
                    return indexed_color_map[indexed_val]
                
                return "#FFFFFF"
            
            # Handle RGB colors
            elif hasattr(color_obj, 'rgb') and color_obj.rgb:
                rgb_value = color_obj.rgb
                if isinstance(rgb_value, str):
                    if len(rgb_value) == 8:
                        return f"#{rgb_value[2:].upper()}"
                    elif len(rgb_value) == 6:
                        return f"#{rgb_value.upper()}"
                return "#FFFFFF"
                
            # Handle theme colors
            elif hasattr(color_obj, 'theme') and color_obj.theme is not None:
                return "#FFFFFF"
                
            # Try indexed access for RGB values
            elif hasattr(color_obj, '__getitem__'):
                try:
                    r, g, b = color_obj[0], color_obj[1], color_obj[2]
                    if isinstance(r, float):
                        r, g, b = int(r * 255), int(g * 255), int(b * 255)
                    return f"#{r:02X}{g:02X}{b:02X}"
                except (IndexError, TypeError, ValueError):
                    pass
                    
            # Try value attribute
            elif hasattr(color_obj, 'value') and color_obj.value:
                rgb_value = color_obj.value
                if isinstance(rgb_value, str):
                    if len(rgb_value) == 8:
                        return f"#{rgb_value[2:].upper()}"
                    elif len(rgb_value) == 6:
                        return f"#{rgb_value.upper()}"
                return "#FFFFFF"
                
            # Direct string conversion
            elif isinstance(color_obj, str):
                if len(color_obj) == 8:
                    return f"#{color_obj[2:].upper()}"
                elif len(color_obj) == 6:
                    return f"#{color_obj.upper()}"
                return "#FFFFFF"
                
            # Try individual RGB attributes
            elif hasattr(color_obj, 'red') and hasattr(color_obj, 'green') and hasattr(color_obj, 'blue'):
                r, g, b = color_obj.red, color_obj.green, color_obj.blue
                if isinstance(r, float):
                    r, g, b = int(r * 255), int(g * 255), int(b * 255)
                return f"#{r:02X}{g:02X}{b:02X}"
                
            else:
                return "#FFFFFF"
            
        except Exception as e:
            self.log(f"Warning: Could not convert color object {type(color_obj)} to hex: {e}")
            return "#FFFFFF"

    def map_color_to_status(self, color: str) -> str:
        """Map color codes to status descriptions"""
        color_mapping = {
            "#FFCC99": "not approved",
            "#CCFFCC": "approved", 
            "#CCFF99": "approved for first order",
            "#FFFFFF": "processing"
        }
        
        return color_mapping.get(color, "unknown")

    def extract_excel_to_json(self, excel_file_path: Path) -> Dict[str, Any]:
        """Extract data and hyperlinks from Excel file using openpyxl"""
        
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        
        start_time = time.time()
        max_row = 5000  # Limit for performance
        
        try:
            self.log("Starting Excel extraction with openpyxl")
            self.log(f"Reading Excel file: {excel_file_path}")
            
            # Load workbook
            load_start = time.time()
            workbook = load_workbook(excel_file_path, data_only=False)
            worksheet = workbook.active
            
            load_time = time.time() - load_start
            self.log(f"Workbook loaded in {load_time:.3f}s")
            
            # Get worksheet dimensions
            actual_max_row = min(worksheet.max_row, max_row)
            max_col = min(worksheet.max_column, 27)  # Limit to AA (27 columns)
            
            self.log(f"Worksheet has {actual_max_row} rows and {max_col} columns")
            
            # Bulk data reading
            data_start = time.time()
            all_data = []
            for row in worksheet.iter_rows(min_row=1, max_row=actual_max_row, max_col=max_col, values_only=True):
                all_data.append(list(row))
            
            data_time = time.time() - data_start
            self.log(f"Bulk data read in {data_time:.3f}s")
            
            # Extract headers from first row
            headers = []
            if all_data:
                header_row = all_data[0]
                for col_idx in range(max_col):
                    if col_idx < len(header_row) and header_row[col_idx]:
                        headers.append(header_row[col_idx])
                    else:
                        headers.append(f"Column_{chr(ord('A') + col_idx)}")
            
            # Identify hyperlink columns (M to U = columns 13 to 21)
            hyperlink_col_indices = list(range(13, min(22, max_col + 1)))
            hyperlink_col_names = [headers[i-1] for i in hyperlink_col_indices if i <= len(headers)]
            self.log(f"Hyperlink columns: {hyperlink_col_names}")
            
            # Extract colors from column A
            color_start = time.time()
            colors = []
            color_count = 0
            
            for row_num in range(2, actual_max_row + 1):
                try:
                    cell = worksheet.cell(row=row_num, column=1)
                    
                    if cell.fill and cell.fill.start_color:
                        hex_color = self.openpyxl_color_to_hex(cell.fill.start_color)
                        
                        if hex_color and hex_color != "#FFFFFF":
                            colors.append(hex_color)
                            color_count += 1
                        else:
                            colors.append("#FFFFFF")
                    else:
                        colors.append("#FFFFFF")
                        
                except Exception as e:
                    self.log(f"Warning: Could not extract color from row {row_num}: {e}")
                    colors.append("#FFFFFF")
            
            color_time = time.time() - color_start
            self.log(f"Colors extracted in {color_time:.3f}s")
            
            # Extract hyperlinks
            hyperlink_start = time.time()
            hyperlink_map = {}
            hyperlink_count = 0
            
            for col_idx in hyperlink_col_indices:
                for row_num in range(2, actual_max_row + 1):
                    try:
                        cell = worksheet.cell(row=row_num, column=col_idx)
                        
                        if cell.hyperlink and cell.hyperlink.target:
                            target = cell.hyperlink.target
                            tooltip = getattr(cell.hyperlink, 'tooltip', None)
                            
                            # Apply URL normalization
                            normalized_target = self.normalize_url(target)
                            
                            hyperlink_map[(row_num, col_idx)] = {
                                'address': target,
                                'subaddress': None,
                                'full_target': normalized_target,
                                'screentip': tooltip
                            }
                            hyperlink_count += 1
                            
                    except Exception:
                        pass
            
            hyperlink_time = time.time() - hyperlink_start
            self.log(f"Hyperlinks extracted in {hyperlink_time:.3f}s ({hyperlink_count} found)")
            
            # Process data rows
            process_start = time.time()
            data_rows = []
            final_hyperlink_count = 0
            final_color_count = 0
            
            for row_idx in range(1, len(all_data)):
                row_num = row_idx + 1
                row_data_tuple = all_data[row_idx]
                row_data = {}
                
                # Extract color and status for this row
                color_idx = row_idx - 1
                if color_idx < len(colors):
                    cell_color = colors[color_idx]
                    if cell_color and cell_color != "#FFFFFF":
                        row_data['color'] = cell_color
                        row_data['status'] = self.map_color_to_status(cell_color)
                        final_color_count += 1
                    else:
                        row_data['color'] = "#FFFFFF"
                        row_data['status'] = "processing"
                else:
                    row_data['color'] = "#FFFFFF"
                    row_data['status'] = "processing"
                
                # Process each column for this row
                for col_idx, header in enumerate(headers):
                    col_num = col_idx + 1
                    
                    if col_num > max_col:
                        break
                    
                    # Get cell value from bulk data
                    cell_value = None
                    if row_data_tuple and col_idx < len(row_data_tuple):
                        cell_value = row_data_tuple[col_idx]
                        
                        # Handle datetime objects
                        if hasattr(cell_value, 'strftime'):
                            cell_value = cell_value.strftime('%Y-%m-%d %H:%M:%S+00:00')
                        
                        # Convert empty strings to None
                        elif isinstance(cell_value, str) and cell_value.strip() == '':
                            cell_value = None
                    
                    # Check if this is a hyperlink column
                    if col_num in hyperlink_col_indices:
                        hyperlink_data = hyperlink_map.get((row_num, col_num))
                        
                        if hyperlink_data:
                            target_url = hyperlink_data['full_target']
                            original_target = hyperlink_data['address']
                            
                            tooltip = hyperlink_data['screentip']
                            if tooltip is None:
                                tooltip = ''
                            
                            row_data[header] = {
                                'display_text': cell_value,
                                'url': target_url,
                                'original_url': None,  # No normalization during extraction
                                'tooltip': tooltip
                            }
                            final_hyperlink_count += 1
                            
                        elif cell_value:
                            # Check if the cell value looks like a file path
                            cell_value_str = str(cell_value).strip()
                            if (cell_value_str.startswith(('C:', 'D:', 'E:', '\\\\', './', '../')) or 
                                '\\' in cell_value_str or 
                                cell_value_str.endswith(('.pdf', '.doc', '.docx', '.xls', '.xlsx'))):
                                
                                # Store raw file path - no normalization during extraction
                                row_data[header] = {
                                    'display_text': cell_value,
                                    'url': cell_value_str,
                                    'original_url': None,
                                    'tooltip': None,
                                    'type': 'inferred_file_path'
                                }
                            else:
                                row_data[header] = {
                                    'display_text': cell_value,
                                    'url': None,
                                    'tooltip': None
                                }
                        else:
                            row_data[header] = None
                    else:
                        # Regular data column
                        if cell_value is None:
                            row_data[header] = None
                        elif isinstance(cell_value, str) and cell_value.strip() == '':
                            row_data[header] = None
                        else:
                            row_data[header] = cell_value
                
                data_rows.append(row_data)
                
                if row_idx % 1000 == 0:
                    self.log(f"  Processed {row_idx} rows...")
            
            process_time = time.time() - process_start
            total_processing_time = time.time() - start_time
            self.log(f"Data processing completed in {process_time:.3f}s")
            
            # Create the final data structure
            data_dict = {
                'metadata': {
                    'total_rows': len(data_rows),
                    'total_columns': len(headers),
                    'columns': headers,
                    'source_file': os.path.basename(excel_file_path),
                    'hyperlinks_extracted': True,
                    'hyperlink_columns': hyperlink_col_names,
                    'colors_extracted': True,
                    'color_mapping': {
                        '#FFCC99': 'not approved',
                        '#CCFFCC': 'approved',
                        '#CCFF99': 'approved for first order',
                        '#FFFFFF': 'processing'
                    },
                    'url_normalization': {
                        'applied': False,
                        'note': 'URL normalization disabled during extraction to prevent conflicts with cleanup step',
                        'normalized_count': 0
                    },
                    'extraction_method': 'openpyxl_ultra_fast',
                    'performance': {
                        'total_time': total_processing_time,
                        'load_time': load_time,
                        'data_read_time': data_time,
                        'color_extraction_time': color_time,
                        'hyperlink_extraction_time': hyperlink_time,
                        'data_processing_time': process_time,
                        'library_used': 'openpyxl',
                        'excel_app_needed': False
                    }
                },
                'data': data_rows
            }
            
            self.log(f"Successfully processed {len(data_rows)} rows")
            self.log(f"Total hyperlinks found: {final_hyperlink_count}")
            self.log(f"Colors extracted: {final_color_count}")
            self.log(f"Total time: {total_processing_time:.3f}s")
            self.log("Note: URL normalization disabled during extraction - will be handled in cleanup step")
            
            return data_dict
            
        except Exception as e:
            self.log(f"Error during Excel extraction: {str(e)}")
            traceback.print_exc()
            raise

    def should_ignore_url(self, url: str) -> bool:
        """Check if a URL should be ignored based on patterns and dead URL list."""
        if not url:
            return True
            
        # Check against ignore patterns
        for pattern in self.ignore_patterns:
            if pattern in url:
                return True
                
        # Check against dead URLs
        if url in self.dead_urls:
            return True
            
        # Check for HTTP/HTTPS URLs (should be ignored)
        if url.startswith(('http://', 'https://')):
            return True
            
        return False

    def fix_url(self, url: str) -> Tuple[str, bool]:
        """Fix a single URL based on replacement rules. Returns (fixed_url, was_changed)"""
        if not url or self.should_ignore_url(url):
            return url, False
            
        original_url = url
        fixed_url = url
        
        # Apply replacement rules
        for old_pattern in self.replacement_rules:
            if old_pattern in fixed_url:
                # Replace the old pattern with the new target
                before_fix = fixed_url
                fixed_url = fixed_url.replace(old_pattern, self.target_replacement)
                
                # Debug: Show rule application for first few fixes
                if self.stats.get('fixed_urls', 0) < 3:
                    self.log(f"    Applied rule '{old_pattern}' -> '{self.target_replacement}'")
                    self.log(f"    Before: {before_fix}")
                    self.log(f"    After:  {fixed_url}")
                
        # Normalize path separators (convert forward slashes to backslashes for Windows paths)
        if fixed_url.startswith('\\\\'):
            fixed_url = fixed_url.replace('/', '\\')
            
        return fixed_url, fixed_url != original_url

    def process_url_object(self, obj: Any) -> bool:
        """Process a URL object (dict with 'url' field) and fix its URL. Returns True if changed."""
        if not isinstance(obj, dict) or 'url' not in obj:
            return False
            
        original_url = obj['url']
        if not original_url:
            return False
            
        self.stats['total_urls'] += 1
        
        # Debug: Show first few URLs being processed
        if self.stats['total_urls'] <= 5:
            self.log(f"  Processing URL #{self.stats['total_urls']}: {original_url}")
        
        if self.should_ignore_url(original_url):
            self.stats['ignored_urls'] += 1
            if self.stats['ignored_urls'] <= 5:
                self.log(f"  -> Ignoring URL (matches ignore pattern): {original_url}")
            return False
            
        fixed_url, was_changed = self.fix_url(original_url)
        
        if was_changed:
            obj['url'] = fixed_url
            self.stats['fixed_urls'] += 1
            self.log(f"  -> Fixed URL: {original_url} -> {fixed_url}")
            return True
        else:
            self.stats['unchanged_urls'] += 1
            if self.stats['unchanged_urls'] <= 5:
                self.log(f"  -> URL unchanged (no matching rules): {original_url}")
            return False

    def process_data_entry(self, entry: Dict, hyperlink_columns: List[str], debug_entry: bool = False) -> int:
        """Process a single data entry and fix all URLs within it. Returns number of changes."""
        changes = 0
        
        for column in hyperlink_columns:
            if column in entry and entry[column]:
                if debug_entry:
                    self.log(f"    Checking column '{column}': {type(entry[column])} = {str(entry[column])[:100]}...")
                
                if self.process_url_object(entry[column]):
                    changes += 1
                    if debug_entry:
                        self.log(f"    -> Fixed URL in column '{column}'")
            elif debug_entry:
                self.log(f"    Column '{column}': Not found or empty")
                    
        return changes

    def cleanup_urls_in_json(self, data: Dict[str, Any]) -> int:
        """Clean up URLs in the JSON data. Returns total number of changes."""
        self.log("Starting URL cleanup...")
        
        # Reset stats
        self.stats = {
            'total_urls': 0,
            'fixed_urls': 0,
            'ignored_urls': 0,
            'unchanged_urls': 0,
            'error_urls': 0
        }
        
        # Get hyperlink columns from metadata or use defaults
        hyperlink_columns = []
        if 'metadata' in data and 'hyperlink_columns' in data['metadata']:
            hyperlink_columns = data['metadata']['hyperlink_columns']
            self.log(f"Using hyperlink columns from metadata: {hyperlink_columns}")
        else:
            # Fallback to hardcoded list
            hyperlink_columns = [
                "Antrag", "Datenblatt", "Produkt-zulassung", "SDB MSDS",
                "Gefährdungsprüfungeurteilung", "Gefährdungsprüfung", 
                "Sonstiges", "Schriftverkehr", "Änd. Historie"
            ]
            self.log(f"Using fallback hyperlink columns: {hyperlink_columns}")
        
        # Debug: Show what we're looking for
        self.log(f"Replacement rules: {self.replacement_rules}")
        self.log(f"Target replacement: {self.target_replacement}")
        
        total_changes = 0
        
        # Process each data entry
        for i, entry in enumerate(data.get('data', [])):
            if i < 3:  # Debug first 3 entries
                self.log(f"DEBUG: Processing entry {i+1}")
                self.log(f"  Available columns: {list(entry.keys())}")
                
            changes = self.process_data_entry(entry, hyperlink_columns, debug_entry=i < 3)
            total_changes += changes
            
            if (i + 1) % 1000 == 0:
                self.log(f"  Processed {i + 1} entries...")
        
        # Update metadata
        if 'metadata' in data:
            data['metadata']['url_cleanup'] = {
                'applied': True,
                'total_changes': total_changes,
                'statistics': self.stats.copy(),
                'rules_applied': len(self.replacement_rules),
                'target_replacement': self.target_replacement,
                'hyperlink_columns_processed': hyperlink_columns
            }
        
        self.log(f"URL cleanup completed:")
        self.log(f"  Total URLs processed: {self.stats['total_urls']}")
        self.log(f"  URLs fixed: {self.stats['fixed_urls']}")
        self.log(f"  URLs ignored: {self.stats['ignored_urls']}")
        self.log(f"  URLs unchanged: {self.stats['unchanged_urls']}")
        self.log(f"  Total changes made: {total_changes}")
        
        return total_changes

    def save_json_data(self, data: Dict[str, Any], temp_path: Path, final_path: Path):
        """Save JSON data to temp location, then move to final location"""
        try:
            # Save to temp location first
            self.log(f"Saving JSON to temp location: {temp_path}")
            with open(temp_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            
            # Create backup of existing file if it exists
            if final_path.exists():
                backup_path = final_path.with_suffix('.json.backup')
                self.log(f"Creating backup: {backup_path}")
                shutil.copy2(final_path, backup_path)
            
            # Move temp file to final location
            self.log(f"Moving to final location: {final_path}")
            shutil.move(str(temp_path), str(final_path))
            
            self.log(f"Successfully saved JSON file: {final_path}")
            
        except Exception as e:
            self.log(f"Error saving JSON: {e}")
            raise

    def cleanup_temp_files(self):
        """Clean up temporary files"""
        try:
            temp_xlsb = self.temp_dir / "Verzeichnis.xlsb"
            temp_xlsx = self.temp_dir / "Verzeichnis.xlsx"
            temp_original_json = self.temp_dir / "Verzeichnis_original_temp.json"
            temp_json = self.temp_dir / "Verzeichnis_temp.json"
            
            for temp_file in [temp_xlsb, temp_xlsx, temp_original_json, temp_json]:
                if temp_file.exists():
                    temp_file.unlink()
                    self.log(f"Deleted temp file: {temp_file}")
                    
        except Exception as e:
            self.log(f"Warning: Could not delete temp files: {e}")

    def run_single_update(self) -> bool:
        """Run a single update cycle. Returns True if successful."""
        try:
            self.log("="*60)
            self.log("STARTING UPDATE CYCLE")
            self.log("="*60)
            
            # Step 1: Fetch source Excel file
            src_path = self.pick_source_path()
            
            # Step 2: Copy to temp directory
            temp_xlsb = self.copy_to_temp_dir(src_path)
            
            # Step 3: Convert XLSB to XLSX
            temp_xlsx = self.convert_xlsb_to_xlsx(temp_xlsb)
            
            # Step 4: Extract data to JSON
            json_data = self.extract_excel_to_json(temp_xlsx)
            
            # Step 5: Save original extracted JSON (before URL cleanup)
            temp_original_json = self.temp_dir / "Verzeichnis_original_temp.json"
            final_original_json = self.data_dir / "Verzeichnis_original.json"
            
            # Create a copy for the original (deep copy to avoid modification)
            import copy
            original_json_data = copy.deepcopy(json_data)
            original_json_data['metadata']['note'] = 'Original extracted data before URL cleanup'
            
            self.save_json_data(original_json_data, temp_original_json, final_original_json)
            self.log(f"Saved original extracted data: {final_original_json}")
            
            # Step 6: Cleanup URLs in the main data
            changes_made = self.cleanup_urls_in_json(json_data)
            
            # Step 7: Save cleaned JSON to final location
            temp_json = self.temp_dir / "Verzeichnis_temp.json"
            final_json = self.data_dir / "Verzeichnis.json"
            self.save_json_data(json_data, temp_json, final_json)
            
            # Step 8: Cleanup temp files
            self.cleanup_temp_files()
            
            self.log("="*60)
            self.log("UPDATE CYCLE COMPLETED SUCCESSFULLY")
            self.log(f"Original file: {final_original_json}")
            self.log(f"Cleaned file: {final_json}")
            self.log(f"Total rows: {json_data['metadata']['total_rows']}")
            self.log(f"URL changes: {changes_made}")
            self.log("="*60)
            
            return True
            
        except Exception as e:
            self.log(f"ERROR in update cycle: {e}")
            traceback.print_exc()
            
            # Try to cleanup temp files even on error
            try:
                self.cleanup_temp_files()
            except:
                pass
                
            return False

    def run_continuous(self, interval_minutes: int = 30):
        """Run continuous updates with specified interval"""
        self.log(f"Starting continuous updates (every {interval_minutes} minutes)")
        self.log("Press Ctrl+C to stop")
        
        try:
            while True:
                success = self.run_single_update()
                
                if success:
                    self.log(f"Waiting {interval_minutes} minutes until next update...")
                else:
                    self.log(f"Update failed, waiting {interval_minutes} minutes before retry...")
                
                # Wait for the specified interval
                time.sleep(interval_minutes * 60)
                
        except KeyboardInterrupt:
            self.log("Continuous updates stopped by user")
        except Exception as e:
            self.log(f"Continuous updates stopped due to error: {e}")
            traceback.print_exc()


def main():
    """Main function"""
    print("Excel Continuous Updater for Django App")
    print("="*50)
    
    # Initialize updater
    try:
        updater = ContinuousExcelUpdater()
    except Exception as e:
        print(f"Failed to initialize updater: {e}")
        return 1
    
    # Check command line arguments
    if len(sys.argv) > 1:
        if sys.argv[1] == "--once":
            # Run single update
            success = updater.run_single_update()
            return 0 if success else 1
        elif sys.argv[1] == "--continuous":
            # Run continuous updates
            interval = 30  # Default 30 minutes
            if len(sys.argv) > 2:
                try:
                    interval = int(sys.argv[2])
                except ValueError:
                    print("Invalid interval, using default 30 minutes")
            
            updater.run_continuous(interval)
            return 0
        else:
            print("Usage:")
            print("  python continuous_updater.py --once")
            print("  python continuous_updater.py --continuous [interval_minutes]")
            return 1
    else:
        # Default: run single update
        success = updater.run_single_update()
        return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())
