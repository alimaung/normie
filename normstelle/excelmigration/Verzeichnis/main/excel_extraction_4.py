import json
import os
import time
from pathlib import Path
from typing import Dict, List, Any, Optional

def normalize_url(url):
    """
    Normalize URLs by replacing relative paths with full network paths
    
    Args:
        url (str): The original URL
        
    Returns:
        str: The normalized URL
    """
    if not url:
        return url
    
    # Replace relative path with full network path
    if url.startswith("../.docs"):
        # Remove "../.docs" and replace with the full network path
        relative_part = url[8:]  # Remove "../.docs" (8 characters)
        # Convert forward slashes to backslashes for Windows network path consistency
        relative_part = relative_part.replace('/', '\\')
        normalized_url = f"file:///\\\\Dehesdna-a009a\\projekte\\k-z\\ofs\\Dokumentenservice\\TeileundStoffe{relative_part}"
        return normalized_url
    
    # If it's already a full file:// URL, return as is
    return url

def rgb_to_hex(rgb_value):
    """
    Convert RGB value to hex color code
    
    Args:
        rgb_value (int/float): RGB value as integer or float
        
    Returns:
        str: Hex color code
    """
    if rgb_value is None:
        return None
    
    try:
        # Convert to integer if it's a float
        rgb_int = int(rgb_value)
        
        # Extract RGB components from the integer
        red = rgb_int & 255
        green = (rgb_int >> 8) & 255
        blue = (rgb_int >> 16) & 255
        
        return f"#{red:02X}{green:02X}{blue:02X}"
    
    except (ValueError, TypeError) as e:
        print(f"Warning: Could not convert RGB value {rgb_value} to hex: {e}")
        return None

def openpyxl_color_to_hex(color_obj):
    """
    Convert openpyxl color object to hex color code
    Handle indexed colors that Excel uses for standard colors
    
    Args:
        color_obj: openpyxl color object (can be RGB object, indexed color, or None)
        
    Returns:
        str: Hex color code
    """
    if not color_obj:
        return "#FFFFFF"
    
    try:
        # CRITICAL: Handle openpyxl indexed colors
        # These are the standard Excel colors that show up as indexed=42, indexed=47, etc.
        if hasattr(color_obj, 'indexed') and color_obj.indexed is not None:
            # Excel indexed color mappings for the colors we care about
            indexed_color_map = {
                42: "#CCFFCC",  # Light green (approved)
                43: "#CCFF99",  # Light green-yellow (approved for first order)  
                47: "#FFCC99",  # Light orange (not approved)
                # Add more indexed colors as needed
            }
            
            indexed_val = color_obj.indexed
            if indexed_val in indexed_color_map:
                return indexed_color_map[indexed_val]
            
            # For unknown indexed colors, default to white
            return "#FFFFFF"
        
        # Handle RGB colors
        elif hasattr(color_obj, 'rgb') and color_obj.rgb:
            rgb_value = color_obj.rgb
            if isinstance(rgb_value, str):
                # String format
                if len(rgb_value) == 8:
                    # ARGB format - remove alpha channel (first 2 chars)
                    return f"#{rgb_value[2:].upper()}"
                elif len(rgb_value) == 6:
                    # RGB format
                    return f"#{rgb_value.upper()}"
            return "#FFFFFF"
            
        # Handle theme colors
        elif hasattr(color_obj, 'theme') and color_obj.theme is not None:
            # For now, default theme colors to white
            # Could be expanded to handle theme color mappings
            return "#FFFFFF"
            
        # Try indexed access for RGB values (safely)
        elif hasattr(color_obj, '__getitem__'):
            try:
                r, g, b = color_obj[0], color_obj[1], color_obj[2]
                # Convert to integers if they're floats (0-1 range)
                if isinstance(r, float):
                    r, g, b = int(r * 255), int(g * 255), int(b * 255)
                return f"#{r:02X}{g:02X}{b:02X}"
            except (IndexError, TypeError, ValueError):
                pass
                
        # Try value attribute
        elif hasattr(color_obj, 'value') and color_obj.value:
            rgb_value = color_obj.value
            if isinstance(rgb_value, str):
                if len(rgb_value) == 8:
                    return f"#{rgb_value[2:].upper()}"
                elif len(rgb_value) == 6:
                    return f"#{rgb_value.upper()}"
            return "#FFFFFF"
            
        # Direct string conversion
        elif isinstance(color_obj, str):
            if len(color_obj) == 8:
                return f"#{color_obj[2:].upper()}"
            elif len(color_obj) == 6:
                return f"#{color_obj.upper()}"
            return "#FFFFFF"
            
        # Try individual RGB attributes
        elif hasattr(color_obj, 'red') and hasattr(color_obj, 'green') and hasattr(color_obj, 'blue'):
            r, g, b = color_obj.red, color_obj.green, color_obj.blue
            # Handle float values (0-1 range)
            if isinstance(r, float):
                r, g, b = int(r * 255), int(g * 255), int(b * 255)
            return f"#{r:02X}{g:02X}{b:02X}"
            
        else:
            return "#FFFFFF"
        
    except Exception as e:
        print(f"Warning: Could not convert color object {type(color_obj)} to hex: {e}")
        return "#FFFFFF"

def map_color_to_status(color):
    """
    Map color codes to status descriptions
    
    Args:
        color (str): Hex color code
        
    Returns:
        str: Status description
    """
    color_mapping = {
        "#FFCC99": "not approved",
        "#CCFFCC": "approved", 
        "#CCFF99": "approved for first order",
        "#FFFFFF": "processing"
    }
    
    return color_mapping.get(color, "unknown")

def extract_excel_to_json_openpyxl(excel_file_path, output_json_path=None, max_row=5000):
    """
    Extract data and hyperlinks from Excel file using openpyxl - ULTRA FAST!
    No Excel application needed, direct XLSX file reading
    
    Args:
        excel_file_path (str): Path to the Excel file
        output_json_path (str, optional): Path for output JSON file
        max_row (int): Maximum row to process (including header)
    
    Returns:
        dict: Extracted data as dictionary
    """
    
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    start_time = time.time()
    
    try:
        print(f"[{time.strftime('%H:%M:%S')}] ============================================================")
        print(f"[{time.strftime('%H:%M:%S')}] OPENPYXL ULTRA-FAST EXCEL EXTRACTION")
        print(f"[{time.strftime('%H:%M:%S')}] ============================================================")
        print(f"[{time.strftime('%H:%M:%S')}] Reading Excel file: {excel_file_path}")
        
        # LOAD WORKBOOK (openpyxl is much faster than COM)
        load_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Loading workbook with openpyxl...")
        
        # Load with data_only=False to get formulas and hyperlinks
        workbook = load_workbook(excel_file_path, data_only=False)
        worksheet = workbook.active
        
        load_time = time.time() - load_start
        print(f"[{time.strftime('%H:%M:%S')}] Workbook loaded in {load_time:.3f}s")
        
        # Get worksheet dimensions
        actual_max_row = min(worksheet.max_row, max_row)
        max_col = min(worksheet.max_column, 27)  # Limit to AA (27 columns)
        
        print(f"[{time.strftime('%H:%M:%S')}] Worksheet has {actual_max_row} rows and {max_col} columns")
        
        # OPTIMIZATION 1: Ultra-fast bulk data reading
        data_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Bulk reading all data...")
        
        # Read all data in one operation using openpyxl's iter_rows
        all_data = []
        for row in worksheet.iter_rows(min_row=1, max_row=actual_max_row, max_col=max_col, values_only=True):
            all_data.append(list(row))
        
        data_time = time.time() - data_start
        print(f"[{time.strftime('%H:%M:%S')}] Bulk data read in {data_time:.3f}s")
        
        # Extract headers from first row
        headers = []
        if all_data:
            header_row = all_data[0]
            for col_idx in range(max_col):
                if col_idx < len(header_row) and header_row[col_idx]:
                    headers.append(header_row[col_idx])
                else:
                    headers.append(f"Column_{chr(ord('A') + col_idx)}")
        
        print(f"[{time.strftime('%H:%M:%S')}] Headers extracted: {headers[:5]}{'...' if len(headers) > 5 else ''}")
        
        # Identify hyperlink columns (M to U = columns 13 to 21)
        hyperlink_col_indices = list(range(13, min(22, max_col + 1)))  # M=13, N=14, ..., U=21
        hyperlink_col_names = [headers[i-1] for i in hyperlink_col_indices if i <= len(headers)]
        print(f"[{time.strftime('%H:%M:%S')}] Hyperlink columns: {hyperlink_col_names}")
        
        # OPTIMIZATION 2: Ultra-fast color extraction
        color_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Extracting colors from column A...")
        
        colors = []
        color_count = 0
        
        for row_num in range(2, actual_max_row + 1):
            try:
                cell = worksheet.cell(row=row_num, column=1)
                
                # Get fill color using openpyxl's native support
                if cell.fill and cell.fill.start_color:
                    hex_color = openpyxl_color_to_hex(cell.fill.start_color)
                    
                    if hex_color and hex_color != "#FFFFFF":
                        colors.append(hex_color)
                        color_count += 1
                        
                        # Debug: Show first few colors
                        if color_count <= 5:
                            print(f"[{time.strftime('%H:%M:%S')}]   Row {row_num}, Column A: Color {hex_color} -> Status: {map_color_to_status(hex_color)}")
                    else:
                        colors.append("#FFFFFF")
                else:
                    colors.append("#FFFFFF")
                    
            except Exception as e:
                print(f"[{time.strftime('%H:%M:%S')}] Warning: Could not extract color from row {row_num}: {e}")
                colors.append("#FFFFFF")
        
        color_time = time.time() - color_start
        print(f"[{time.strftime('%H:%M:%S')}] Colors extracted in {color_time:.3f}s")
        
        # OPTIMIZATION 3: Ultra-fast hyperlink extraction
        hyperlink_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Extracting hyperlinks from target columns...")
        
        hyperlink_map = {}
        hyperlink_count = 0
        
        # Only check hyperlink columns
        for col_idx in hyperlink_col_indices:
            for row_num in range(2, actual_max_row + 1):
                try:
                    cell = worksheet.cell(row=row_num, column=col_idx)
                    
                    # Check if cell has hyperlink using openpyxl's native support
                    if cell.hyperlink and cell.hyperlink.target:
                        target = cell.hyperlink.target
                        tooltip = getattr(cell.hyperlink, 'tooltip', None)
                        
                        # CRITICAL FIX: Apply URL normalization here
                        normalized_target = normalize_url(target)
                        
                        hyperlink_map[(row_num, col_idx)] = {
                            'address': target,
                            'subaddress': None,  # openpyxl handles this differently
                            'full_target': normalized_target,  # Use normalized target
                            'screentip': tooltip
                        }
                        hyperlink_count += 1
                        
                        # Debug: Show first few hyperlinks
                        if hyperlink_count <= 5:
                            print(f"[{time.strftime('%H:%M:%S')}]   Row {row_num}, Col {col_idx}: '{cell.value}' -> {normalized_target[:50]}...")
                            
                except Exception as e:
                    # Skip cells that cause errors
                    pass
        
        hyperlink_time = time.time() - hyperlink_start
        print(f"[{time.strftime('%H:%M:%S')}] Hyperlinks extracted in {hyperlink_time:.3f}s")
        print(f"[{time.strftime('%H:%M:%S')}] Found {hyperlink_count} hyperlinks")
        
        # OPTIMIZATION 4: Ultra-fast in-memory data processing
        process_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Processing {len(all_data) - 1} rows in memory...")
        
        data_rows = []
        normalized_count = 0
        final_hyperlink_count = 0
        final_color_count = 0
        
        # Process data rows (skip header row)
        for row_idx in range(1, len(all_data)):
            row_num = row_idx + 1  # Excel row number
            row_data_tuple = all_data[row_idx]
            row_data = {}
            
            # Extract color and status for this row
            color_idx = row_idx - 1  # Adjust for colors array (which starts from row 2)
            if color_idx < len(colors):
                cell_color = colors[color_idx]
                if cell_color and cell_color != "#FFFFFF":
                    row_data['color'] = cell_color
                    row_data['status'] = map_color_to_status(cell_color)
                    final_color_count += 1
                else:
                    row_data['color'] = "#FFFFFF"
                    row_data['status'] = "processing"
            else:
                row_data['color'] = "#FFFFFF"
                row_data['status'] = "processing"
            
            # Process each column for this row
            for col_idx, header in enumerate(headers):
                col_num = col_idx + 1  # Convert to 1-based column number
                
                if col_num > max_col:
                    break
                
                # Get cell value from bulk data with proper None/empty handling
                cell_value = None
                if row_data_tuple and col_idx < len(row_data_tuple):
                    cell_value = row_data_tuple[col_idx]
                    
                    # Handle datetime objects to match COM format (add timezone info)
                    if hasattr(cell_value, 'strftime'):  # datetime object
                        cell_value = cell_value.strftime('%Y-%m-%d %H:%M:%S+00:00')
                    
                    # Convert empty strings to None for consistency, but preserve actual text
                    elif isinstance(cell_value, str) and cell_value.strip() == '':
                        cell_value = None
                
                # Check if this is a hyperlink column
                if col_num in hyperlink_col_indices:
                    # Check if there's a hyperlink for this cell
                    hyperlink_data = hyperlink_map.get((row_num, col_num))
                    
                    if hyperlink_data:
                        # Process hyperlink - the normalization was already done in hyperlink extraction
                        normalized_target = hyperlink_data['full_target']
                        original_target = hyperlink_data['address']
                        
                        if original_target != normalized_target:
                            normalized_count += 1
                        
                        # Handle tooltip properly - convert None to empty string to match COM behavior
                        tooltip = hyperlink_data['screentip']
                        if tooltip is None:
                            tooltip = ''
                        
                        row_data[header] = {
                            'display_text': cell_value,
                            'url': normalized_target,
                            'original_url': original_target if original_target != normalized_target else None,
                            'tooltip': tooltip
                        }
                        final_hyperlink_count += 1
                        
                    elif cell_value:
                        # Check if the cell value looks like a file path
                        cell_value_str = str(cell_value).strip()
                        if (cell_value_str.startswith(('C:', 'D:', 'E:', '\\\\', './', '../')) or 
                            '\\' in cell_value_str or 
                            cell_value_str.endswith(('.pdf', '.doc', '.docx', '.xls', '.xlsx'))):
                            
                            # Normalize the inferred file path
                            normalized_path = normalize_url(cell_value_str)
                            
                            row_data[header] = {
                                'display_text': cell_value,
                                'url': normalized_path,
                                'original_url': cell_value_str if cell_value_str != normalized_path else None,
                                'tooltip': None,
                                'type': 'inferred_file_path'
                            }
                            
                            if cell_value_str != normalized_path:
                                normalized_count += 1
                        else:
                            # Regular text value in hyperlink column
                            row_data[header] = {
                                'display_text': cell_value,
                                'url': None,
                                'tooltip': None
                            }
                    else:
                        # Empty cell in hyperlink column
                        row_data[header] = None
                else:
                    # Regular data column - ensure consistent None vs empty string handling
                    if cell_value is None:
                        row_data[header] = None
                    elif isinstance(cell_value, str) and cell_value.strip() == '':
                        row_data[header] = None  # Convert empty strings to None
                    else:
                        row_data[header] = cell_value
            
            data_rows.append(row_data)
            
            # Progress indicator for large datasets
            if row_idx % 1000 == 0:
                print(f"[{time.strftime('%H:%M:%S')}]   Processed {row_idx} rows...")
        
        process_time = time.time() - process_start
        print(f"[{time.strftime('%H:%M:%S')}] Data processing completed in {process_time:.3f}s")
        
        # Create the final data structure
        total_processing_time = time.time() - start_time
        
        data_dict = {
            'metadata': {
                'total_rows': len(data_rows),
                'total_columns': len(headers),
                'columns': headers,
                'source_file': os.path.basename(excel_file_path),
                'hyperlinks_extracted': True,
                'hyperlink_columns': hyperlink_col_names,
                'colors_extracted': True,
                'color_mapping': {
                    '#FFCC99': 'not approved',
                    '#CCFFCC': 'approved',
                    '#CCFF99': 'approved for first order',
                    '#FFFFFF': 'processing'
                },
                'url_normalization': {
                    'applied': True,
                    'rule': 'Replace "../.docs" with "file:///\\\\Dehesdna-a009a\\projekte\\k-z\\ofs\\Dokumentenservice\\TeileundStoffe"',
                    'normalized_count': normalized_count
                },
                'extraction_method': 'openpyxl_ultra_fast',
                'performance': {
                    'total_time': total_processing_time,
                    'load_time': load_time,
                    'data_read_time': data_time,
                    'color_extraction_time': color_time,
                    'hyperlink_extraction_time': hyperlink_time,
                    'data_processing_time': process_time,
                    'library_used': 'openpyxl',
                    'excel_app_needed': False
                }
            },
            'data': data_rows
        }
        
        print(f"[{time.strftime('%H:%M:%S')}] Successfully processed {len(data_rows)} rows")
        print(f"[{time.strftime('%H:%M:%S')}] Total hyperlinks found: {final_hyperlink_count}")
        print(f"[{time.strftime('%H:%M:%S')}] URLs normalized: {normalized_count}")
        print(f"[{time.strftime('%H:%M:%S')}] Colors extracted: {final_color_count}")
        
        # Performance breakdown
        print(f"\n[{time.strftime('%H:%M:%S')}] ============================================================")
        print(f"[{time.strftime('%H:%M:%S')}] OPENPYXL PERFORMANCE BREAKDOWN:")
        print(f"[{time.strftime('%H:%M:%S')}] ============================================================")
        print(f"[{time.strftime('%H:%M:%S')}] Workbook loading: {load_time:.3f}s ({load_time/total_processing_time*100:.1f}%)")
        print(f"[{time.strftime('%H:%M:%S')}] Bulk data reading: {data_time:.3f}s ({data_time/total_processing_time*100:.1f}%)")
        print(f"[{time.strftime('%H:%M:%S')}] Color extraction: {color_time:.3f}s ({color_time/total_processing_time*100:.1f}%)")
        print(f"[{time.strftime('%H:%M:%S')}] Hyperlink extraction: {hyperlink_time:.3f}s ({hyperlink_time/total_processing_time*100:.1f}%)")
        print(f"[{time.strftime('%H:%M:%S')}] Data processing: {process_time:.3f}s ({process_time/total_processing_time*100:.1f}%)")
        print(f"[{time.strftime('%H:%M:%S')}] TOTAL TIME: {total_processing_time:.3f}s")
        
        # Handle None values for JSON serialization
        def clean_none_values(obj):
            if isinstance(obj, dict):
                return {k: clean_none_values(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [clean_none_values(item) for item in obj]
            elif obj is None:
                return None
            else:
                return obj
        
        data_dict = clean_none_values(data_dict)
        
        # Determine output file path
        if output_json_path is None:
            excel_path = Path(excel_file_path)
            output_json_path = excel_path.parent / f"{excel_path.stem}_openpyxl.json"
        
        # Write to JSON file
        write_start = time.time()
        print(f"[{time.strftime('%H:%M:%S')}] Writing JSON to: {output_json_path}")
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(data_dict, f, indent=2, ensure_ascii=False, default=str)
        
        write_time = time.time() - write_start
        total_time = time.time() - start_time
        
        print(f"[{time.strftime('%H:%M:%S')}] JSON written in {write_time:.3f}s")
        print(f"[{time.strftime('%H:%M:%S')}] Successfully created openpyxl JSON file with {len(data_rows)} rows")
        print(f"[{time.strftime('%H:%M:%S')}] Output file: {output_json_path}")
        
        # Final performance summary
        print(f"\n[{time.strftime('%H:%M:%S')}] ============================================================")
        print(f"[{time.strftime('%H:%M:%S')}] FINAL PERFORMANCE SUMMARY")
        print(f"[{time.strftime('%H:%M:%S')}] ============================================================")
        print(f"[{time.strftime('%H:%M:%S')}] TOTAL EXECUTION TIME: {total_time:.3f}s")
        
        # Compare with previous versions
        original_time = 492.8  # Original script
        v2_time = 154.7       # Our best pywin32 optimization
        
        print(f"[{time.strftime('%H:%M:%S')}] vs Original script: {original_time/total_time:.1f}x FASTER ({original_time:.1f}s → {total_time:.1f}s)")
        print(f"[{time.strftime('%H:%M:%S')}] vs Our best v2: {v2_time/total_time:.1f}x FASTER ({v2_time:.1f}s → {total_time:.1f}s)")
        print(f"[{time.strftime('%H:%M:%S')}] Library: openpyxl (no Excel application needed)")
        print(f"[{time.strftime('%H:%M:%S')}] ============================================================")
        
        return data_dict
        
    except Exception as e:
        print(f"Error during openpyxl extraction: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

def main():
    """Main function to run the openpyxl ultra-fast extraction"""
    # Get the directory of this script
    script_dir = Path(__file__).parent
    excel_file = script_dir / "Verzeichnis.xlsx"
    
    if not excel_file.exists():
        print(f"Error: Excel file not found at {excel_file}")
        return
    
    try:
        start_time = time.time()
        
        # Extract data and hyperlinks using openpyxl
        data = extract_excel_to_json_openpyxl(str(excel_file))
        
        total_time = time.time() - start_time
        
        # Print summary
        print(f"\n[{time.strftime('%H:%M:%S')}] ============================================================")
        print(f"[{time.strftime('%H:%M:%S')}] OPENPYXL ULTRA-FAST EXTRACTION SUMMARY")
        print(f"[{time.strftime('%H:%M:%S')}] ============================================================")
        print(f"[{time.strftime('%H:%M:%S')}] Source file: {excel_file.name}")
        print(f"[{time.strftime('%H:%M:%S')}] Total rows extracted: {data['metadata']['total_rows']}")
        print(f"[{time.strftime('%H:%M:%S')}] Total columns: {data['metadata']['total_columns']}")
        print(f"[{time.strftime('%H:%M:%S')}] Hyperlink columns: {', '.join(data['metadata']['hyperlink_columns'])}")
        print(f"[{time.strftime('%H:%M:%S')}] URLs normalized: {data['metadata']['url_normalization']['normalized_count']}")
        print(f"[{time.strftime('%H:%M:%S')}] Colors extracted: {data['metadata'].get('colors_extracted', False)}")
        print(f"[{time.strftime('%H:%M:%S')}] Extraction method: {data['metadata']['extraction_method']}")
        print(f"[{time.strftime('%H:%M:%S')}] Excel app needed: {data['metadata']['performance']['excel_app_needed']}")
        print(f"[{time.strftime('%H:%M:%S')}] TOTAL TIME: {total_time:.3f}s")
        
        # Show performance data
        if 'performance' in data['metadata']:
            perf = data['metadata']['performance']
            print(f"[{time.strftime('%H:%M:%S')}] Performance breakdown:")
            print(f"[{time.strftime('%H:%M:%S')}]   Library: {perf['library_used']}")
            print(f"[{time.strftime('%H:%M:%S')}]   Load time: {perf['load_time']:.3f}s")
            print(f"[{time.strftime('%H:%M:%S')}]   Data read: {perf['data_read_time']:.3f}s")
            print(f"[{time.strftime('%H:%M:%S')}]   Colors: {perf['color_extraction_time']:.3f}s")
            print(f"[{time.strftime('%H:%M:%S')}]   Hyperlinks: {perf['hyperlink_extraction_time']:.3f}s")
        
        # Show sample data
        if data['data']:
            print(f"\n[{time.strftime('%H:%M:%S')}] First row sample:")
            first_row = data['data'][0]
            
            # Show color and status
            print(f"[{time.strftime('%H:%M:%S')}]   Color: {first_row.get('color', 'None')}")
            print(f"[{time.strftime('%H:%M:%S')}]   Status: {first_row.get('status', 'None')}")
            
            # Show regular columns
            regular_cols = ['Antrag-nummer', 'Teile-nummer', 'Freigabe']
            for col in regular_cols:
                if col in first_row:
                    print(f"[{time.strftime('%H:%M:%S')}]   {col}: {first_row[col]}")
            
            # Show hyperlink columns
            print(f"\n[{time.strftime('%H:%M:%S')}] Hyperlink columns sample:")
            for col in data['metadata']['hyperlink_columns'][:3]:
                if col in first_row and first_row[col]:
                    if isinstance(first_row[col], dict):
                        print(f"[{time.strftime('%H:%M:%S')}]   {col}: {first_row[col]['display_text']} -> {first_row[col]['url'][:50]}...")
                    else:
                        print(f"[{time.strftime('%H:%M:%S')}]   {col}: {first_row[col]}")
                        
    except Exception as e:
        print(f"Failed to extract data: {str(e)}")

if __name__ == "__main__":
    main()
